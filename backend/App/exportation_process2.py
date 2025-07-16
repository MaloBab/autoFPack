import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from App import models

logger = logging.getLogger("fpack_export")
logger.setLevel(logging.DEBUG)
if not logger.handlers:
    handler = logging.FileHandler("export_fpack.log", mode="w", encoding="utf-8")
    handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(handler)


def get_item_label(type_: str, ref_id: int, db: Session):
    """
    Retourne (label principal, détail) pour un item donné
    """
    try:
        if type_ == "produit":
            prod = db.query(models.Produit).filter_by(id=ref_id).first()
            if prod:
                return prod.nom, prod.description or ""
            return f"Produit {ref_id}", ""
        elif type_ == "equipement":
            return f"Équipement {ref_id}", ""
        elif type_ == "robot":
            rob = db.query(models.Robots).filter_by(id=ref_id).first()
            if rob:
                return rob.nom, rob.generation or ""
            return f"Robot {ref_id}", ""
        else:
            return f"{type_} {ref_id}", ""
    except Exception as e:
        logger.error(f"Erreur get_item_label type={type_} ref_id={ref_id}: {e}")
        return f"{type_} {ref_id}", ""


def export_fpack_to_sheet(fpack_id: int, ws, db: Session):
    """
    Écrit trois tableaux verticaux dans la feuille, séparés de 2 lignes.
    """
    # Styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, color="FFFFFF")
    orange_fill = PatternFill("solid", fgColor="FF9933")
    green_fill = PatternFill("solid", fgColor="70AD47")
    blue_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center")

    fpack = db.query(models.FPack).filter_by(id=fpack_id).first()
    config_columns = db.query(models.FPackConfigColumn).filter_by(fpack_id=fpack_id).order_by(models.FPackConfigColumn.ordre).all()

    # === En-tête général ===
    ws.merge_cells("A1:H1")
    ws["A1"] = "F-Pack Matrix"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["A2"] = f"Nom: {fpack.nom}"
    ws["C2"] = f"Abbr: {fpack.fpack_abbr}"
    ws["E2"] = "Généré automatiquement"

    current_row = 4  # Ligne de départ pour le premier tableau

    # -------------------------
    # GROUPES
    # -------------------------
    groupes = [col for col in config_columns if col.type == "group"]
    if groupes:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell = ws.cell(row=current_row, column=1, value="GROUPES")
        cell.font = header_font
        cell.fill = orange_fill
        cell.alignment = center

        current_row += 1
        for col_cfg in groupes:
            groupe = db.query(models.Groupes).filter_by(id=col_cfg.ref_id).first()
            items = db.query(models.GroupeItem).filter_by(group_id=col_cfg.ref_id).all()
            labels, details = zip(*[get_item_label(item.type, item.ref_id, db) for item in items])

            # Header
            ws.cell(row=current_row, column=1, value=groupe.nom if groupe else f"Groupe {col_cfg.ref_id}").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Détails").font = Font(italic=True)

            # Données
            for i, (label, detail) in enumerate(zip(labels, details), start=1):
                ws.cell(row=current_row+i, column=1, value=label)
                ws.cell(row=current_row+i, column=2, value=detail)

            # Avancer la ligne après le dernier item
            current_row += len(labels) + 1

        current_row += 2  # 2 lignes vides


    # -------------------------
    # ÉQUIPEMENTS
    # -------------------------
    equipements = [col for col in config_columns if col.type == "equipement"]
    if equipements:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        cell = ws.cell(row=current_row, column=1, value="ÉQUIPEMENTS")
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = center

        current_row += 1
        for col_cfg in equipements:
            eq = db.query(models.Equipements).filter_by(id=col_cfg.ref_id).first()

            # Header
            ws.cell(row=current_row, column=1, value=eq.nom if eq else f"Équipement {col_cfg.ref_id}").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Détails").font = Font(italic=True)

            # Exemple : remplir avec description (si dispo)
            ws.cell(row=current_row+1, column=1, value=eq.nom if eq else f"Équipement {col_cfg.ref_id}")
            ws.cell(row=current_row+1, column=2, value=detail)

            current_row += 2  # passer à la prochaine paire

        current_row += 2  # 2 lignes vides


    # -------------------------
    # PRODUITS
    # -------------------------
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    cell = ws.cell(row=current_row, column=1, value="PRODUITS")
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center

    headers = ["Nom", "Fournisseur", "Type", "Description"]
    current_row += 1
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=current_row, column=idx, value=header)
        c.font = Font(bold=True)
        c.fill = blue_fill
        c.alignment = center

    produits = db.query(models.Produit).all()
    for i, prod in enumerate(produits, start=1):
        fournisseur = db.query(models.Fournisseur).filter_by(id=prod.fournisseur_id).first()
        ws.cell(row=current_row+i, column=1, value=prod.nom)
        ws.cell(row=current_row+i, column=2, value=fournisseur.nom if fournisseur else "")
        ws.cell(row=current_row+i, column=3, value=prod.type)
        ws.cell(row=current_row+i, column=4, value=prod.description)


def export_fpack_config(fpack_id: int, db: Session) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "FPack"
    export_fpack_to_sheet(fpack_id, ws, db)
    return wb


def export_all_fpacks(db: Session) -> Workbook:
    wb = Workbook()
    fpacks = db.query(models.FPack).all()
    for i, fpack in enumerate(fpacks):
        ws = wb.active if i == 0 else wb.create_sheet(title=fpack.nom or f"FPack_{fpack.id}")
        export_fpack_to_sheet(fpack.id, ws, db)
    return wb
