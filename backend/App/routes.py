from fastapi import APIRouter, Depends, HTTPException 
from sqlalchemy.orm import Session 
from App.database import SessionLocal
from App import models, schemas
from sqlalchemy import inspect 
from sqlalchemy.orm import selectinload
from App.export_fpack_to_excel import export_fpack_config, export_all_fpacks
from fastapi.responses import StreamingResponse
from io import BytesIO
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

os.makedirs("logs", exist_ok=True)

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/table-columns/{table_name}", response_model=list[str])
def get_table_columns(table_name: str, db: Session = Depends(get_db)):
    inspector = inspect(db.bind)

    use_sql_server = os.getenv("USE_SQL_SERVER", "false").lower() == "true"

    if use_sql_server:
        table_mapping = {
            "clients": "FPM_clients",
            "produits": "FPM_produits",
            "equipements": "FPM_equipements",
            "robots": "FPM_robots",
            "fournisseurs": "FPM_fournisseurs",
            "fpacks": "FPM_fpacks",
            "prix": "FPM_prix",
            "prix_robot": "FPM_prix_robot",
            "projets": "FPM_projets",
        }

        actual_name = table_mapping.get(table_name)
        if not actual_name:
            raise HTTPException(status_code=404, detail=f"Table inconnue : {table_name}")
    else:
        actual_name = table_name

    columns = inspector.get_columns(actual_name)
    return [col["name"] for col in columns]


# PRODUITS
@router.get("/produits", response_model=list[schemas.ProduitRead])
def list_produits(db: Session = Depends(get_db)):
    return db.query(models.Produit).all()

@router.post("/produits", response_model=schemas.ProduitRead)
def create_produit(produit: schemas.ProduitCreate, db: Session = Depends(get_db)):
    db_produit = models.Produit(**produit.dict())
    db.add(db_produit)
    db.commit()
    db.refresh(db_produit)
    return db_produit

@router.put("/produits/{id}", response_model=schemas.ProduitRead)
def update_produit(id: int, produit: schemas.ProduitCreate, db: Session = Depends(get_db)):
    db_produit = db.query(models.Produit).get(id)
    if not db_produit:
        raise HTTPException(status_code=404, detail="Produit non trouvé")
    for key, value in produit.dict().items():
        setattr(db_produit, key, value)
    db.commit()
    db.refresh(db_produit)
    return db_produit

@router.delete("/produits/{id}")
def delete_produit(id: int, db: Session = Depends(get_db)):
    db_produit = db.query(models.Produit).get(id)
    if not db_produit:
        raise HTTPException(status_code=404, detail="Produit non trouvé")

    equipements = db.query(models.Equipement_Produit).filter(
        models.Equipement_Produit.produit_id == id
    ).all()

    if equipements:
        equipement_ids = {ep.equipement_id for ep in equipements}
        noms = db.query(models.Equipements.nom).filter(models.Equipements.id.in_(equipement_ids)).all()
        noms_liste = [nom for (nom,) in noms]
        raise HTTPException(
            status_code=400,
            detail=f"Suppression impossible : {db_produit.nom} est lié aux équipement(s) : {', '.join(noms_liste)}"
        )

    db.delete(db_produit)
    db.commit()
    return {"ok": True}

@router.get("/produits/{id}", response_model=schemas.ProduitRead)
def get_produit(id: int, db: Session = Depends(get_db)):
    produit = db.query(models.Produit).get(id)
    if not produit:
        raise HTTPException(status_code=404, detail="Produit non trouvé")
    return produit


@router.post("/produits/{produit_id}/duplicate", response_model=schemas.ProduitRead)
def duplicate_produit(produit_id: int, db: Session = Depends(get_db)):
    original = db.query(models.Produit).filter(models.Produit.id == produit_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Produit non trouvé")

    new_produit = models.Produit(
        nom=original.nom + " (copie)",
        description=original.description,
        fournisseur_id=original.fournisseur_id,
        type=original.type,
    )
    db.add(new_produit)
    db.commit()
    db.refresh(new_produit)

    return new_produit

# PRIX

@router.get("/prix", response_model=list[schemas.PrixRead])
def list_prix(db: Session = Depends(get_db)):
    return db.query(models.Prix).all()

@router.post("/prix", response_model=schemas.PrixRead)
def create_prix(prix: schemas.PrixCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Prix).filter(
        models.Prix.produit_id == prix.produit_id,
        models.Prix.client_id == prix.client_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Prix déjà défini pour ce produit et client")

    db_prix = models.Prix(**prix.dict())
    db.add(db_prix)
    db.commit()
    db.refresh(db_prix)
    return db_prix

@router.put("/prix/{produit_id}/{client_id}", response_model=schemas.PrixRead)
def update_prix(produit_id: int, client_id: int, prix: schemas.PrixCreate, db: Session = Depends(get_db)):
    db_prix = db.query(models.Prix).filter(
        models.Prix.produit_id == produit_id,
        models.Prix.client_id == client_id
    ).first()
    if not db_prix:
        raise HTTPException(status_code=404, detail="Prix non trouvé")
    for key, value in prix.model_dump().items():
        setattr(db_prix, key, value)
    db.commit()
    db.refresh(db_prix)
    return db_prix

@router.delete("/prix/{produit_id}/{client_id}")
def delete_prix(produit_id: int, client_id: int, db: Session = Depends(get_db)):
    db_prix = db.query(models.Prix).filter(
        models.Prix.produit_id == produit_id,
        models.Prix.client_id == client_id
    ).first()
    if not db_prix:
        raise HTTPException(status_code=404, detail="Prix non trouvé")
    db.delete(db_prix)
    db.commit()
    return {"ok": True}

@router.get("/prix/{produit_id}/{client_id}", response_model=schemas.PrixRead)
def get_prix(produit_id: int, client_id: int, db: Session = Depends(get_db)):
    db_prix = db.query(models.Prix).filter(
        models.Prix.produit_id == produit_id,
        models.Prix.client_id == client_id
    ).first()
    if not db_prix:
        raise HTTPException(status_code=404, detail="Prix non trouvé")
    return db_prix



# FOURNISSEURS
@router.get("/fournisseurs", response_model=list[schemas.FournisseurRead])
def list_fournisseurs(db: Session = Depends(get_db)):
    return db.query(models.Fournisseur).all()

@router.post("/fournisseurs", response_model=schemas.FournisseurRead)
def create_fournisseur(fournisseur: schemas.FournisseurCreate, db: Session = Depends(get_db)):
    db_fournisseur = models.Fournisseur(**fournisseur.dict())
    db.add(db_fournisseur)
    db.commit()
    db.refresh(db_fournisseur)
    return db_fournisseur


@router.put("/fournisseurs/{id}", response_model=schemas.FournisseurRead)
def update_fournisseur(id: int, fournisseur: schemas.FournisseurCreate, db: Session = Depends(get_db)):
    db_fournisseur = db.query(models.Fournisseur).get(id)
    if not db_fournisseur:
        raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
    for key, value in fournisseur.dict().items():
        setattr(db_fournisseur, key, value)
    db.commit()
    db.refresh(db_fournisseur)
    return db_fournisseur


@router.delete("/fournisseurs/{id}")
def delete_fournisseur(id: int, db: Session = Depends(get_db)):
    db_fournisseur = db.query(models.Fournisseur).get(id)
    if not db_fournisseur:
        raise HTTPException(status_code=404, detail="Fournisseur non trouvé")
    db.delete(db_fournisseur)
    db.commit()
    return {"ok": True}

# CLIENTS
@router.get("/clients", response_model=list[schemas.ClientRead])
def list_clients(db: Session = Depends(get_db)):
    return db.query(models.Client).all()

@router.post("/clients", response_model=schemas.ClientRead)
def create_client(client: schemas.ClientCreate, db: Session = Depends(get_db)):
    db_client = models.Client(**client.dict())
    db.add(db_client)
    db.commit()
    db.refresh(db_client)
    return db_client

@router.put("/clients/{id}", response_model=schemas.ClientRead)
def update_client(id: int, client: schemas.ClientCreate, db: Session = Depends(get_db)):
    db_client = db.query(models.Client).get(id)
    if not db_client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    for key, value in client.dict().items():
        setattr(db_client, key, value)
    db.commit()
    db.refresh(db_client)
    return db_client

@router.delete("/clients/{id}")
def delete_client(id: int, db: Session = Depends(get_db)):
    db_client = db.query(models.Client).get(id)
    if not db_client:
        raise HTTPException(status_code=404, detail="Client non trouvé")

    # Vérifie si ce client est utilisé dans des robots
    robots = db.query(models.Robots).filter(models.Robots.client == id).all()
    fpacks = db.query(models.FPack).filter(models.FPack.client == id).all()

    noms_robots = [r.nom for r in robots]
    noms_fpacks = [f.nom for f in fpacks]

    erreurs = []
    if noms_robots:
        erreurs.append(f"robots : {', '.join(noms_robots)}")
    if noms_fpacks:
        erreurs.append(f"fpacks : {', '.join(noms_fpacks)}")

    if erreurs:
        raise HTTPException(
            status_code=400,
            detail=f"Suppression impossible : {db_client.nom} est lié aux {', et aux '.join(erreurs)}"
        )

    db.delete(db_client)
    db.commit()
    return {"ok": True}

# ROBOTS
@router.get("/robots", response_model=list[schemas.RobotRead])
def list_robots(db: Session = Depends(get_db)):
    return db.query(models.Robots).all()

@router.post("/robots", response_model=schemas.RobotRead)
def create_robot(robot: schemas.RobotCreate, db: Session = Depends(get_db)):
    db_robot = models.Robots(**robot.dict())
    db.add(db_robot)
    db.commit()
    db.refresh(db_robot)
    return db_robot

@router.put("/robots/{id}", response_model=schemas.RobotRead)
def update_robot(id: int, robot: schemas.RobotCreate, db: Session = Depends(get_db)):
    db_robot = db.query(models.Robots).get(id)
    if not db_robot:
        raise HTTPException(status_code=404, detail="Robot non trouvé")
    for key, value in robot.dict().items():
        setattr(db_robot, key, value)
    db.commit()
    db.refresh(db_robot)
    return db_robot

@router.delete("/robots/{id}")
def delete_robot(id: int, db: Session = Depends(get_db)):
    db_robot = db.query(models.Robots).get(id)
    if not db_robot:
        raise HTTPException(status_code=404, detail="Robot non trouvé")
    db.delete(db_robot)
    db.commit()
    return {"ok": True}

# EQUIPEMENTS
@router.get("/equipements", response_model=list[schemas.EquipementRead])
def list_equipements(db: Session = Depends(get_db)):
    return db.query(models.Equipements).all()

@router.post("/equipements", response_model=schemas.EquipementRead)
def create_equipement(equipement: schemas.EquipementCreate, db: Session = Depends(get_db)):
    db_equipement = models.Equipements(**equipement.dict())
    db.add(db_equipement)
    db.commit()
    db.refresh(db_equipement)
    return db_equipement

@router.put("/equipements/{id}", response_model=schemas.EquipementRead)
def update_equipement(id: int, equipement: schemas.EquipementCreate, db: Session = Depends(get_db)):
    db_equipement = db.query(models.Equipements).get(id)
    if not db_equipement:
        raise HTTPException(status_code=404, detail="Equipement non trouvé")
    for key, value in equipement.dict().items():
        setattr(db_equipement, key, value)
    db.commit()
    db.refresh(db_equipement)
    return db_equipement

@router.delete("/equipements/{id}")
def delete_equipement(id: int, db: Session = Depends(get_db)):
    db_equipement = db.query(models.Equipements).get(id)
    if not db_equipement:
        raise HTTPException(status_code=404, detail="Equipement non trouvé")
    db.delete(db_equipement)
    db.commit()
    return {"ok": True}


@router.get("/equipements/{id}", response_model=schemas.EquipementRead)
def get_equipement(id: int, db: Session = Depends(get_db)):
    equipement = db.query(models.Equipements)\
        .options(selectinload(models.Equipements.equipement_produit))\
        .filter(models.Equipements.id == id)\
        .first()
    if not equipement:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    return equipement

# EQUIPEMENTS PRODUITS

@router.get("/equipementproduits", response_model=dict[int, list[schemas.EquipementProduitRead]])
def get_all_equipement_produits(db: Session = Depends(get_db)):
    results = db.query(models.Equipement_Produit).all()

    equipement_dict: dict[int, list[models.Equipement_Produit]] = {}
    for ep in results:
        equipement_dict.setdefault(ep.equipement_id, []).append(ep)

    return equipement_dict

@router.get("/equipementproduit/{equipement_id}", response_model=list[schemas.EquipementProduitRead])
def get_equipement_produit_by_equipement(equipement_id: int, db: Session = Depends(get_db)):
    return db.query(models.Equipement_Produit).filter(models.Equipement_Produit.equipement_id == equipement_id).all()

@router.post("/equipementproduit", response_model=schemas.EquipementProduitRead)
def create_or_update_equipement_produit(equipement_produit: schemas.EquipementProduitCreate, db: Session = Depends(get_db)):
    ep = db.query(models.Equipement_Produit).filter(
        models.Equipement_Produit.equipement_id == equipement_produit.equipement_id,
        models.Equipement_Produit.produit_id == equipement_produit.produit_id
    ).first()
    if ep:
        ep.quantite = equipement_produit.quantite
    else:
        ep = models.Equipement_Produit(**equipement_produit.dict())
    db.add(ep)

    db.commit()
    db.refresh(ep)
    return ep


@router.delete("/equipementproduit/{id}")
def delete_equipement_produit(id: int, db: Session = Depends(get_db)):
    db_equipement_produit = db.query(models.Equipement_Produit).get(id)
    if not db_equipement_produit:
        raise HTTPException(status_code=404, detail="Equipement de produit non trouvé")
    db.delete(db_equipement_produit)
    db.commit()
    return {"ok": True}

@router.delete("/equipementproduit/clear/{equipement_id}")
def clear_equipement_produits(equipement_id: int, db: Session = Depends(get_db)):
    db.query(models.Equipement_Produit).filter(models.Equipement_Produit.equipement_id == equipement_id).delete()
    db.commit()
    return {"ok": True}


@router.get("/equipements/{equipement_id}")
def get_equipement(equipement_id: int, db: Session = Depends(get_db)):
    equipement = db.query(models.Equipements).filter(models.Equipements.id == equipement_id).first()
    if not equipement:
        raise HTTPException(status_code=404, detail="Equipement not found")
    return equipement


# FPACKS
@router.get("/fpacks", response_model=list[schemas.FPackRead])
def list_fpacks(db: Session = Depends(get_db)):
    return db.query(models.FPack).all()

@router.post("/fpacks", response_model=schemas.FPackRead)
def create_fpack(fpack: schemas.FPackCreate, db: Session = Depends(get_db)):
    db_fpack = models.FPack(**fpack.dict())
    db.add(db_fpack)
    db.commit()
    db.refresh(db_fpack)
    return db_fpack

@router.put("/fpacks/{id}", response_model=schemas.FPackRead)
def update_fpack(id: int, fpack: schemas.FPackCreate, db: Session = Depends(get_db)):
    db_fpack = db.query(models.FPack).get(id)
    if not db_fpack:
        raise HTTPException(status_code=404, detail="FPack non trouvé")
    for key, value in fpack.dict().items():
        setattr(db_fpack, key, value)
    db.commit()
    db.refresh(db_fpack)
    return db_fpack

@router.delete("/fpacks/{id}")
def delete_fpack(id: int, db: Session = Depends(get_db)):
    db_fpack = db.query(models.FPack).get(id)
    if not db_fpack:
        raise HTTPException(status_code=404, detail="FPack non trouvé")
    db.delete(db_fpack)
    db.commit()
    return {"ok": True}

@router.get("/fpacks/{id}", response_model=schemas.FPackRead)
def get_fpack(id: int, db: Session = Depends(get_db)):
    db_fpack = db.query(models.FPack).get(id)
    if not db_fpack:
        raise HTTPException(status_code=404, detail="FPack non trouvé")
    return db_fpack

@router.post("/fpacks/{fpack_id}/duplicate", response_model=schemas.FPackRead)
def duplicate_fpack(fpack_id: int, db: Session = Depends(get_db)):
    original_fpack = db.query(models.FPack).filter(models.FPack.id == fpack_id).first()
    if not original_fpack:
        raise HTTPException(status_code=404, detail="FPack non trouvée")

    new_fpack = models.FPack(
        nom=original_fpack.nom + " (copie)",
        client=original_fpack.client,
        fpack_abbr=original_fpack.fpack_abbr
    )
    db.add(new_fpack)
    db.commit()
    db.refresh(new_fpack)

    config_columns = db.query(models.FPackConfigColumn).filter_by(fpack_id=fpack_id).all()
    for col in config_columns:
        copied_col = models.FPackConfigColumn(
            fpack_id=new_fpack.id,
            ordre=col.ordre,
            type=col.type,
            ref_id=col.ref_id
        )
        db.add(copied_col)

    db.commit()

    return new_fpack


# GROUPS
@router.get("/groupes", response_model=list[schemas.GroupesRead])
def list_groupes(db: Session = Depends(get_db)):
    return db.query(models.Groupes).all()

@router.post("/groupes", response_model=schemas.GroupesRead)
def create_groupe(group: schemas.GroupesCreate, db: Session = Depends(get_db)):
    db_group = models.Groupes(**group.dict())
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    return db_group

@router.get("/groupe_items/{groupe_id}", response_model=list[schemas.GroupeItemRead])
def list_groupe_items(groupe_id: int, db: Session = Depends(get_db)):
    return db.query(models.GroupeItem).filter(models.GroupeItem.group_id == groupe_id).all()


@router.post("/groupe_items", response_model=schemas.GroupeItemRead)
def create_groupe_item(item: schemas.GroupeItemCreate, db: Session = Depends(get_db)):
    db_item = models.GroupeItem(**item.dict())
    db.add(db_item)
    db.commit()
    db.refresh(db_item)

    return db_item

# FPACK CONFIG COLUMNS
@router.get("/fpack_config_columns/{fpack_id}")
def get_config_columns(fpack_id: int, db: Session = Depends(get_db)):
    columns = db.query(models.FPackConfigColumn).filter_by(fpack_id=fpack_id).order_by(models.FPackConfigColumn.ordre).all()
    result = []

    for col in columns:
        entry = {
            "type": col.type,
            "ref_id": col.ref_id,
            "ordre": col.ordre
        }

        if col.type == "produit":
            produit = db.query(models.Produit).filter_by(id=col.ref_id).first()
            entry["display_name"] = produit.nom if produit else f"(produit {col.ref_id})"

        elif col.type == "equipement":
            eq = db.query(models.Equipements).filter_by(id=col.ref_id).first()
            entry["display_name"] = eq.nom if eq else f"(équipement {col.ref_id})"

        elif col.type == "group":
            groupe = db.query(models.Groupes).filter_by(id=col.ref_id).first()
            entry["display_name"] = groupe.nom if groupe else f"(groupe {col.ref_id})"

            items = db.query(models.GroupeItem).filter_by(group_id=col.ref_id).all()
            entry["group_items"] = [
                {
                    "type": item.type,
                    "ref_id": item.ref_id,
                    "label": get_item_label(item.type, item.ref_id, db),
                    "statut": item.statut,
                    
                }
                for item in items
            ]

        result.append(entry)
        
        with open("logs/fpack_config_columns.log", "a") as log_file:
            log_file.write(f"{result}\n")

    return result


def get_item_label(type_: str, ref_id: int, db: Session = Depends(get_db)):
    if type_ == "produit":
        item = db.query(models.Produit).filter_by(id=ref_id).first()
        return item.nom if item else f"(produit {ref_id})"
    elif type_ == "equipement":
        item = db.query(models.Equipements).filter_by(id=ref_id).first()
        return item.nom if item else f"(équipement {ref_id})"
    elif type_ == "robot":
        item = db.query(models.Robots).filter_by(id=ref_id).first()
        return item.nom if item else f"(robot {ref_id})"
    return f"{type_} {ref_id}"

@router.post("/fpack_config_columns", response_model=schemas.FPackConfigColumnRead)
def create_fpack_column(col: schemas.FPackConfigColumnCreate, db: Session = Depends(get_db)):
    db_col = models.FPackConfigColumn(**col.dict())
    db.add(db_col)
    db.commit()
    db.refresh(db_col)
    return db_col

@router.put("/fpack_config_columns/{id}", response_model=schemas.FPackConfigColumnRead)
def update_fpack_column(id: int, col: schemas.FPackConfigColumnCreate, db: Session = Depends(get_db)):
    db_col = db.query(models.FPackConfigColumn).get(id)
    if not db_col:
        raise HTTPException(status_code=404, detail="Colonne de configuration non trouvée")
    for key, value in col.dict().items():
        setattr(db_col, key, value)
    db.commit()
    db.refresh(db_col)
    return db_col

@router.delete("/fpack_config_columns/{id}")
def delete_fpack_column(id: int, db: Session = Depends(get_db)):
    db_col = db.query(models.FPackConfigColumn).get(id)
    if not db_col:
        raise HTTPException(status_code=404, detail="Colonne de configuration non trouvée")
    db.delete(db_col)
    db.commit()
    return {"ok": True}

@router.delete("/fpack_config_columns/clear/{fpack_id}")
def clear_fpack_config_columns(fpack_id: int, db: Session = Depends(get_db)):
    db.query(models.FPackConfigColumn).filter(models.FPackConfigColumn.fpack_id == fpack_id).delete()
    db.commit()
    return {"ok": True}

@router.get("/dashboard/stats")
def dashboard_stats(db: Session = Depends(get_db)):
    return {
        "produits": db.query(models.Produit).count(),
        "equipements": db.query(models.Equipements).count(),
        "robots": db.query(models.Robots).count(),
        "clients": db.query(models.Client).count(),
        "fournisseurs": db.query(models.Fournisseur).count(),
        "fpacks": db.query(models.FPack).count()
    }
    
# PRODUIT-INCOMPATIBILITÉS
@router.get("/produit-incompatibilites", response_model=list[schemas.ProduitIncompatibiliteRead])
def list_incompatibilites(db: Session = Depends(get_db)):
    return db.query(models.ProduitIncompatibilite).all()

@router.post("/produit-incompatibilites", response_model=schemas.ProduitIncompatibiliteRead)
def create_incompatibilite(incomp: schemas.ProduitIncompatibiliteCreate, db: Session = Depends(get_db)):
    db_incomp = models.ProduitIncompatibilite(**incomp.dict())
    db.add(db_incomp)
    db.commit()
    return db_incomp


# ROBOT-PRODUIT-INCOMPATIBILITÉS
@router.get("/robot-produit-incompatibilites", response_model=list[schemas.RobotProduitIncompatibiliteRead])
def list_robot_incompatibilites(db: Session = Depends(get_db)):
    return db.query(models.RobotProduitIncompatibilite).all()

@router.post("/robot-produit-incompatibilites", response_model=schemas.RobotProduitIncompatibiliteRead)
def create_robot_incompatibilite(incomp: schemas.RobotProduitIncompatibiliteCreate, db: Session = Depends(get_db)):
    db_incomp = models.RobotProduitIncompatibilite(**incomp.dict())
    db.add(db_incomp)
    db.commit()
    return db_incomp

@router.delete("/produit-incompatibilites")
def delete_produit_incompatibilite(inc: schemas.ProduitIncompatibiliteCreate, db: Session = Depends(get_db)):
    db.query(models.ProduitIncompatibilite).filter(
        models.ProduitIncompatibilite.produit_id_1 == inc.produit_id_1,
        models.ProduitIncompatibilite.produit_id_2 == inc.produit_id_2
    ).delete()
    db.commit()
    return {"ok": True}

@router.delete("/robot-produit-incompatibilites")
def delete_robot_produit_incompatibilite(incomp: schemas.RobotProduitIncompatibiliteCreate, db: Session = Depends(get_db)):
    db.query(models.RobotProduitIncompatibilite).filter(
        models.RobotProduitIncompatibilite.robot_id == incomp.robot_id,
        models.RobotProduitIncompatibilite.produit_id == incomp.produit_id
    ).delete()
    db.commit()
    return {"ok": True}

#EXPORT FPACK TO EXCEL
@router.post("/export-fpack/{fpack_id}")
def export_fpack(fpack_id: int, db: Session = Depends(get_db)):
    wb = export_fpack_config(fpack_id, db)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        'Content-Disposition': f'attachment; filename="F-Pack-{fpack_id}.xlsx"'
    }

    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.get("/export-fpacks/all")
def export_all_fpacks_route(db: Session = Depends(get_db)):
    wb = export_all_fpacks(db)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename="FPacks-All.xlsx"'
    }

    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

#PROJET

@router.post("/projets", response_model=schemas.ProjetRead)
def create_projet(projet: schemas.ProjetCreate, db: Session = Depends(get_db)):
    db_projet = models.Projet(nom=projet.nom, client=projet.client, fpack_id=projet.fpack_id)
    db.add(db_projet)
    db.commit()
    db.refresh(db_projet)
    return db_projet

@router.get("/projets", response_model=list[schemas.ProjetRead])
def list_projets(db: Session = Depends(get_db)):
    return db.query(models.Projet).all()

@router.get("/projets/{id}", response_model=schemas.ProjetRead)
def get_projet(id: int, db: Session = Depends(get_db)):
    projet = db.query(models.Projet).filter_by(id=id).first()
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    return projet

@router.delete("/projets/{id}")
def delete_projet(id: int, db: Session = Depends(get_db)):
    db_projet = db.query(models.Projet).filter_by(id=id).first()
    if not db_projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    db.delete(db_projet)
    db.commit()
    return {"ok": True}

@router.put("/projets/{id}", response_model=schemas.ProjetRead)
def update_projet(id: int, projet: schemas.ProjetCreate, db: Session = Depends(get_db)):
    db_projet = db.query(models.Projet).filter_by(id=id).first()
    if not db_projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Met à jour les champs de base
    db_projet.nom = projet.nom
    db_projet.client = projet.client
    db_projet.fpack_id = projet.fpack_id
    db.commit()
    db.refresh(db_projet)
    return db_projet

#############

@router.get("/projets/{id}/selections", response_model=list[schemas.ProjetSelectionRead])
def get_projet_selections(id: int, db: Session = Depends(get_db)):
    return db.query(models.ProjetSelection).filter(models.ProjetSelection.projet_id == id).all()

from fastapi import Body

@router.put("/projets/{id}/selections", response_model=dict)
def save_projet_selections(
    id: int,
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    db.query(models.ProjetSelection).filter(models.ProjetSelection.projet_id == id).delete()
    db.commit()
    for sel in data.get("selections", []):
        selection = models.ProjetSelection(
            projet_id=id,
            groupe_id=sel.get("groupe_id"),
            ref_id=sel.get("ref_id"),
            type_item=sel.get("type_item", "produit")  # ou "equipement" ou "groupe" selon le contexte
        )
        db.add(selection)
    db.commit()
    return {"message": "Sélections enregistrées"}

### FACTURE

@router.get("/projets/{id}/facture")
def get_projet_facture(id: int, db: Session = Depends(get_db)):
    from collections import defaultdict

    # 1. Charger le projet
    projet = db.query(models.Projet).filter_by(id=id).first()
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    client_id = projet.client
    fpack_id = projet.fpack_id

    # 2. Config du FPack
    config_cols = db.query(models.FPackConfigColumn).filter(models.FPackConfigColumn.fpack_id == fpack_id).order_by(models.FPackConfigColumn.ordre).all()

    # 3. Sélections projet
    sels = db.query(models.ProjetSelection).filter(models.ProjetSelection.projet_id == id).all()
    sel_map = {s.groupe_id: s.ref_id for s in sels}

    # 4. Mapping équipements -> produits
    eq_prods = db.query(models.Equipement_Produit).all()
    eq_map: dict[int, list[int]] = {}
    for ep in eq_prods:
        eq_map.setdefault(ep.equipement_id, []).append((ep.produit_id, ep.quantite))

    # 5. Compter les produits (avec quantités)
    produit_counts = defaultdict(int)
    robot_counts = defaultdict(int)

    # produits seuls
    for col in config_cols:
        if col.type == "produit":
            produit_counts[col.ref_id] += 1

    # équipements seuls
    for col in config_cols:
        if col.type == "equipement":
            for pid in eq_map.get(col.ref_id, []):
                produit_counts[pid] += 1

    # groupes
    for col in config_cols:
        if col.type == "group":
            chosen_ref = sel_map.get(col.ref_id)
            if not chosen_ref:
                continue
            gi = db.query(models.GroupeItem).filter_by(group_id=col.ref_id, ref_id=chosen_ref).first()
            if not gi:
                continue
            if gi.type == "produit":
                produit_counts[gi.ref_id] += 1
            elif gi.type == "equipement":
                for pid, qte in eq_map.get(gi.ref_id, []):  # correction ici
                    produit_counts[pid] += qte
            elif gi.type == "robot":
                robot_counts[gi.ref_id] += 1  # à prix 0

    all_produit_ids = list(produit_counts.keys())

    # 6. Charger prix
    prix_rows = db.query(models.Prix)\
        .filter(models.Prix.client_id == client_id, models.Prix.produit_id.in_(all_produit_ids))\
        .all()
    prix_map = { (p.produit_id, p.client_id): p for p in prix_rows }

    # 7. Charger noms produits
    produits_rows = db.query(models.Produit).filter(models.Produit.id.in_(all_produit_ids)).all()
    produit_nom_map = {p.id: p.nom for p in produits_rows}

    # 8. Construire les lignes
    lines = []
    total_produit = 0
    total_transport = 0

    for pid in sorted(all_produit_ids):
        qte = produit_counts[pid]
        p_rec = prix_map.get((pid, client_id))
        prix_prod = float(p_rec.prix_produit) if p_rec else 0.0
        prix_tr = float(p_rec.prix_transport) if p_rec else 0.0
        commentaire = p_rec.commentaire if p_rec else None

        total_ligne = qte * (prix_prod + prix_tr)
        total_produit += qte * prix_prod
        total_transport += qte * prix_tr

        lines.append({
            "produit_id": pid,
            "nom": produit_nom_map.get(pid, f"Produit {pid}"),
            "qte": qte,
            "prix_produit": prix_prod,
            "prix_transport": prix_tr,
            "commentaire": commentaire,
            "total_ligne": total_ligne
        })
    """
    # 9. Ajouter les robots
    robot_rows = db.query(models.Robot).filter(models.Robot.id.in_(robot_counts.keys())).all()
    robot_nom_map = {r.id: r.nom for r in robot_rows}

    for rid in sorted(robot_counts.keys()):
        qte = robot_counts[rid]
        lines.append({
            "produit_id": rid,
            "nom": robot_nom_map.get(rid, f"Robot {rid}"),
            "qte": qte,
            "prix_produit": 0.0,
            "prix_transport": 0.0,
            "commentaire": "Robot (non facturé -- temporaire)",
            "total_ligne": 0.0
        })
    """

    return {
        "projet_id": id,
        "nom_projet": projet.nom,
        "client_id": client_id,
        "fpack_id": fpack_id,
        "currency": "EUR",
        "lines": lines,
        "totaux": {
            "produit": total_produit,
            "transport": total_transport,
            "global": total_produit + total_transport
        }
    }
    
@router.get("/projets/{id}/facture-pdf")
def export_projet_facture_pdf(id: int, db: Session = Depends(get_db)):
    facture = get_projet_facture(id, db)
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)

    elements = []
    styles = getSampleStyleSheet()

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    logo_path = os.path.join(BASE_DIR, "frontend", "assets", "FANUCLOGO.jpg")
    if os.path.exists(logo_path):
        img = Image(logo_path, width=100, height=40)
        elements.append(img)

    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>FACTURE - {facture['nom_projet']}</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    data = [["Produit", "Quantité", "Prix Unité", "Transport", "Commentaire", "Total ligne"]]
    for ligne in facture["lines"]:
        data.append([
            ligne.get("nom", ""),
            ligne.get("qte", 0),
            f"{ligne.get('prix_produit', 0)} €",
            f"{ligne.get('prix_transport', 0)} €",
            ligne.get("commentaire", ""),
            f"{ligne.get('total_ligne', 0)} €"
        ])

    table = Table(data, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.yellow),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.red),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 24))

    for label, value in facture["totaux"].items():
        elements.append(Paragraph(f"<b>{label.title()} :</b> {value} €", styles['Normal']))

    doc.build(elements)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="facture-projet-{id}.pdf"'}
    )

@router.get("/projets/{id}/facture-excel")
def export_projet_facture_excel(id: int, db: Session = Depends(get_db)):
    facture = get_projet_facture(id, db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Facture Projet"

    ws.merge_cells("C2:F2")
    cell = ws["C2"]
    cell.value = f"FACTURE - {facture['nom_projet']}"
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal="center")

    headers = ["Produit", "Quantité", "Prix Unité", "Transport", "Commentaire", "Total ligne"]
    ws.append([""] * len(headers))  
    ws.append(headers)
    header_row = ws.max_row
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="AA0000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ligne in facture["lines"]:
        ws.append([
            ligne.get("nom", ""),
            ligne.get("qte", 0),
            float(ligne.get("prix_produit") or 0),
            float(ligne.get("prix_transport") or 0),
            ligne.get("commentaire", ""),
            float(ligne.get("total_ligne") or 0)
        ])

    ws.append([])
    ws.append(["", "", "", "", "Total produit", facture["totaux"]["produit"]])
    ws.append(["", "", "", "", "Total transport", facture["totaux"]["transport"]])
    ws.append(["", "", "", "", "Total global", facture["totaux"]["global"]])

    for i in range(1, 8):
        ws.column_dimensions[chr(64 + i)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="facture-projet-{id}.xlsx"'}
    )
