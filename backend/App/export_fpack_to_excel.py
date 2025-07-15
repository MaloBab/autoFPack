import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session
from App import models

# === Configuration du logger ===
logger = logging.getLogger("fpack_export")
logger.setLevel(logging.DEBUG)

file_handler = logging.FileHandler("export_fpack.log", mode="w", encoding="utf-8")
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

if not logger.hasHandlers():
    logger.addHandler(file_handler)


# === Utilitaire pour récupérer le nom lisible d’un item ===
def get_item_label(type_: str, ref_id: int, db: Session) -> str:
    try:
        if type_ == "produit":
            item = db.query(models.Produit).filter_by(id=ref_id).first()
            logger.debug(f"[Produit] #{ref_id} → {item.nom if item else 'introuvable'}")
            return item.nom if item else f"Produit {ref_id}"
        elif type_ == "equipement":
            item = db.query(models.Equipements).filter_by(id=ref_id).first()
            logger.debug(f"[Équipement] #{ref_id} → {item.nom if item else 'introuvable'}")
            return item.nom if item else f"Équipement {ref_id}"
        elif type_ == "robot":
            item = db.query(models.Robots).filter_by(id=ref_id).first()
            logger.debug(f"[Robot] #{ref_id} → {item.nom if item else 'introuvable'}")
            return item.nom if item else f"Robot {ref_id}"
        else:
            logger.warning(f"[Inconnu] type={type_} ref_id={ref_id}")
            return f"{type_} {ref_id}"
    except Exception as e:
        logger.error(f"[ERREUR get_item_label] type={type_}, ref_id={ref_id} → {e}")
        return f"{type_} {ref_id}"


# === Fonction principale d’export ===
def export_fpack_config(fpack_id: int, db: Session) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "F-Pack"


    ws["A1"].value = "F-Pack Matrix"
    fpack = db.query(models.FPack).filter(models.FPack.id == fpack_id).first()
    ws["A2"].value = f"Nom: {fpack.nom}"
    ws["A3"].value = f"Abbr: {fpack.fpack_abbr}"
    ws["A4"].value = "Généré automatiquement"

    logger.info(f"Export de la FPack #{fpack_id} - {fpack.nom} ({fpack.fpack_abbr})")

    config_columns = (
        db.query(models.FPackConfigColumn)
        .filter(models.FPackConfigColumn.fpack_id == fpack_id)
        .order_by(models.FPackConfigColumn.ordre)
        .all()
    )
    logger.debug(f"Nombre de colonnes dans la configuration : {len(config_columns)}")
    
    
    # === Fusion et en-têtes ligne 1 à 4 ===
    for row in range(1, 5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column= len(config_columns) + 3)
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")

    col_index = 1
    nb_col_groupes = 0

    # === Colonnes de type group ===
    for col in config_columns:
        if col.type != "group":
            continue

        groupe = db.query(models.Groupes).filter_by(id=col.ref_id).first()
        group_items = db.query(models.GroupeItem).filter_by(group_id=col.ref_id).all()
        logger.debug(f"[Group] {col.ref_id} → {groupe.nom if groupe else 'introuvable'} avec {len(group_items)} items")

        options = [get_item_label(item.type, item.ref_id, db) for item in group_items]
        options = [opt for opt in options if opt]
        options = options[:30]
        logger.debug(f"→ Options affichées : {options}")

        header = ws.cell(row=5, column=col_index)
        header.value = groupe.nom if groupe else f"Groupe {col.ref_id}"
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal="center")
        header.fill = PatternFill("solid", fgColor="FFCC99")

        formula_str = ",".join(options)
        logger.debug(f"Longueur formule validation : {len(formula_str)} caractères")

        dv = DataValidation(type="list", formula1=f'"{formula_str}"')
        ws.add_data_validation(dv)
        row = 6
        dv.add(ws.cell(row=row, column=col_index))
        logger.debug(f"→ Validation ajoutée colonne {col_index}")

        ws.column_dimensions[chr(64 + col_index)].width = 24
        col_index += 1
        nb_col_groupes += 1

    produits_seuls = []
    equipements_seuls = []

    for col in config_columns:
        if col.type == "produit":
            item = db.query(models.Produit).filter_by(id=col.ref_id).first()
            produits_seuls.append(item.nom if item else f"Produit {col.ref_id}")
        elif col.type == "equipement":
            item = db.query(models.Equipements).filter_by(id=col.ref_id).first()
            equipements_seuls.append(item.nom if item else f"Équipement {col.ref_id}")

    # === Section à part à droite du tableau ===
    ligne_depart = 5
    col_equipements = nb_col_groupes + 3
    col_produits = nb_col_groupes + 4


    # En-tête Produits
    header = ws.cell(row=ligne_depart, column=col_produits, value="Produits")
    header.font = Font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="4472C4")
    for i, nom in enumerate(produits_seuls, start=1):
        ws.cell(row=ligne_depart + i, column=col_produits, value=nom)

    # En-tête Équipements
    header = ws.cell(row=ligne_depart, column=col_equipements, value="Équipements")
    header.font = Font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="70AD47")
    for i, nom in enumerate(equipements_seuls, start=1):
        ws.cell(row=ligne_depart + i, column=col_equipements, value=nom)

    ws.column_dimensions[chr(64 + col_equipements)].width = 24
    ws.column_dimensions[chr(64 + col_equipements+ 1)].width = 24
    
    ws.freeze_panes = "A6"
    logger.info("Export terminé ✅")
    return wb
