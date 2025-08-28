from App import models
from App.database import SessionLocal
from fastapi import APIRouter, HTTPException, Depends #type: ignore
from sqlalchemy.orm import Session #type: ignore
from typing import List, Dict, Any, Optional
import pandas as pd #type: ignore
from io import BytesIO
import traceback
from openpyxl import Workbook #type: ignore
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment #type: ignore
from fastapi.responses import StreamingResponse #type: ignore

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

class FPackExportService:
    """Service d'export des F-Packs avec formatage Excel"""
    
    def __init__(self, db: Session):
        self.db = db
        
    def get_fpack_data_for_export(self, sous_projet_fpack_ids: List[int]) -> List[Dict[str, Any]]:
        """Récupère les données des instances F-Pack (SousProjetFpack) pour l'export"""
        try:
            fpacks_query = self.db.query(models.SousProjetFpack)\
                .join(models.SousProjet, models.SousProjetFpack.sous_projet_id == models.SousProjet.id)\
                .join(models.ProjetGlobal, models.SousProjet.id_global == models.ProjetGlobal.id)\
                .join(models.Client, models.ProjetGlobal.client == models.Client.id)\
                .join(models.FPack, models.SousProjetFpack.fpack_id == models.FPack.id)\
                .filter(models.SousProjetFpack.id.in_(sous_projet_fpack_ids))\
                .all()
            
            if not fpacks_query:
                raise HTTPException(
                    status_code=404, 
                    detail="Aucune instance F-Pack trouvée avec les IDs fournis"
                )
            
            export_data = []
            
            for sous_projet_fpack in fpacks_query:
                selections = self._get_fpack_selections(sous_projet_fpack.id)
                produits_seuls, equipements_seuls = self._get_fpack_standalone_items(sous_projet_fpack.fpack_id)
                fpack_data = self._build_base_fpack_data(sous_projet_fpack)
                fpack_data.update(self._build_selections_data(selections))
                fpack_data.update(self._build_standalone_items_data(produits_seuls, equipements_seuls))
                
                export_data.append(fpack_data)
            
            return export_data
            
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Erreur lors de la récupération des données: {str(e)}"
            )
            
    def _get_fpack_standalone_items(self, fpack_id: int) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Récupère les produits et équipements seuls de la configuration F-Pack"""
        config_columns = self.db.query(models.FPackConfigColumn)\
            .filter(
                models.FPackConfigColumn.fpack_id == fpack_id,
                models.FPackConfigColumn.type.in_(['produit', 'equipement'])
            )\
            .order_by(models.FPackConfigColumn.ordre)\
            .all()
        
        produits_seuls = []
        equipements_seuls = []
        
        for col in config_columns:
            if col.type == 'produit':
                produit = self.db.query(models.Produit).filter(models.Produit.id == col.ref_id).first()
                if produit:
                    produits_seuls.append({
                        'id': produit.id,
                        'nom': produit.nom,
                        'reference': getattr(produit, 'reference', ''),
                        'ordre': col.ordre
                    })
            elif col.type == 'equipement':
                equipement = self.db.query(models.Equipements).filter(models.Equipements.id == col.ref_id).first()
                if equipement:
                    equipements_seuls.append({
                        'id': equipement.id,
                        'nom': equipement.nom,
                        'reference': getattr(equipement, 'reference', ''),
                        'ordre': col.ordre
                    })
        
        return produits_seuls, equipements_seuls
    
    def _get_fpack_selections(self, sous_projet_fpack_id: int) -> List[Dict[str, Any]]:
        """Récupère les sélections d'une instance F-Pack"""
        selections_query = self.db.query(
            models.ProjetSelection.groupe_id,
            models.Groupes.nom.label('groupe_nom'),
            models.ProjetSelection.type_item,
            models.ProjetSelection.ref_id
        )\
        .join(models.Groupes, models.ProjetSelection.groupe_id == models.Groupes.id)\
        .filter(models.ProjetSelection.sous_projet_fpack_id == sous_projet_fpack_id)\
        .all()
        
        selections = []
        for selection in selections_query:
            item_nom = self._get_item_name(selection.type_item, selection.ref_id)
            
            selections.append({
                'groupe_id': selection.groupe_id,
                'groupe_nom': selection.groupe_nom,
                'type_item': selection.type_item,
                'ref_id': selection.ref_id,
                'item_nom': item_nom
            })
        
        return selections
    
    def _get_item_name(self, type_item: str, ref_id: int) -> str:
        """Récupère le nom d'un item selon son type"""
        try:
            if type_item == 'produit':
                item = self.db.query(models.Produit.nom).filter(models.Produit.id == ref_id).first()
                return item.nom if item else f"Produit {ref_id}"
            elif type_item == 'equipement':
                item = self.db.query(models.Equipements.nom).filter(models.Equipements.id == ref_id).first()
                return item.nom if item else f"Equipement {ref_id}"
            elif type_item == 'robot':
                item = self.db.query(models.Robots.nom).filter(models.Robots.id == ref_id).first()
                return item.nom if item else f"Robot {ref_id}"
            else:
                return f"Item {ref_id}"
        except Exception as e:
            return f"{type_item} {ref_id}"
    
    def _build_base_fpack_data(self, sous_projet_fpack) -> Dict[str, Any]:
        """Construit les données de base de l'instance F-Pack"""
        return {
            'Projet': sous_projet_fpack.sous_projet.global_rel.projet,
            'Client': sous_projet_fpack.sous_projet.global_rel.client_rel.nom,
            'Sous_Projet': sous_projet_fpack.sous_projet.nom,
            
            'FPack_Number': sous_projet_fpack.FPack_number or '',
            'Robot_Location_Code': sous_projet_fpack.Robot_Location_Code or '',
            'Contractor': sous_projet_fpack.contractor or '',
            'Required_Delivery_Time': sous_projet_fpack.required_delivery_time or '',
            'Delivery_Site': sous_projet_fpack.delivery_site or '',
            'Tracking': sous_projet_fpack.tracking or '',
            
            'FPack_Template': sous_projet_fpack.fpack.nom or '',
            'FPack_Abbreviation': sous_projet_fpack.fpack.fpack_abbr or '',
        }
    
    def _build_base_fpack_data(self, sous_projet_fpack) -> Dict[str, Any]:
        """Construit les données de base de l'instance F-Pack"""
        return {
            'Projet': sous_projet_fpack.sous_projet.global_rel.projet,
            'Client': sous_projet_fpack.sous_projet.global_rel.client_rel.nom,
            'Sous_Projet': sous_projet_fpack.sous_projet.nom,
            
            'FPack_Number': sous_projet_fpack.FPack_number or '',
            'Robot_Location_Code': sous_projet_fpack.Robot_Location_Code or '',
            'Contractor': sous_projet_fpack.contractor or '',
            'Required_Delivery_Time': sous_projet_fpack.required_delivery_time or '',
            'Delivery_Site': sous_projet_fpack.delivery_site or '',
            'Tracking': sous_projet_fpack.tracking or '',
            
            'FPack_Template': sous_projet_fpack.fpack.nom or '',
            'FPack_Abbreviation': sous_projet_fpack.fpack.fpack_abbr or '',
        }
    
    def _build_selections_data(self, selections: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Construit les données des sélections par groupe"""
        selections_data = {}
        
        for selection in selections:
            groupe_nom = selection['groupe_nom']
            item_nom = selection['item_nom']
            
            if groupe_nom not in selections_data:
                selections_data[groupe_nom] = []
            
            selections_data[groupe_nom].append(item_nom)
            
        for groupe_nom, items in selections_data.items():
            selections_data[groupe_nom] = ', '.join(items)
        
        return selections_data
    
    def _build_standalone_items_data(self, produits_seuls: List[Dict[str, Any]], equipements_seuls: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Construit les données des produits et équipements seuls"""
        standalone_data = {}
        
        if produits_seuls:
            produits_noms = [p['nom'] for p in sorted(produits_seuls, key=lambda x: x['ordre'])]
            standalone_data['Produits_Seuls'] = ', '.join(produits_noms)
        else:
            standalone_data['Produits_Seuls'] = ''
        
        if equipements_seuls:
            equipements_noms = [e['nom'] for e in sorted(equipements_seuls, key=lambda x: x['ordre'])]
            standalone_data['Equipements_Seuls'] = ', '.join(equipements_noms)
        else:
            standalone_data['Equipements_Seuls'] = ''
        
        return standalone_data
    
    def create_excel_export(self, export_data: List[Dict[str, Any]]) -> BytesIO:
        """Crée le fichier Excel formaté selon le modèle F-Pack Matrix"""
        try:
            df = pd.DataFrame(export_data)
            
            if df.empty:
                raise ValueError("Aucune donnée à exporter")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "F-Pack Matrix"
            
            self._setup_excel_styles(wb, ws)
            self._add_headers_and_data(ws, df)
            self._apply_formatting(ws, df)
            self._auto_adjust_columns(ws)
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Erreur lors de la génération du fichier Excel: {str(e)}"
            )
    
    def _setup_excel_styles(self, wb: Workbook, ws):
        """Configure les styles Excel"""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        self.header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        self.sub_header_font = Font(name="Arial", size=9, bold=True, color="000000")
        self.data_font = Font(name="Arial", size=9, color="000000")
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def _add_headers_and_data(self, ws, df: pd.DataFrame):
        """Ajoute les en-têtes et données au worksheet"""
        headers = list(df.columns)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = self.center_alignment
        
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = self.left_alignment

                if row_idx % 2 == 0:
                    cell.fill = self.data_fill
    
    def _apply_formatting(self, ws, df: pd.DataFrame):
        """Applique le formatage spécifique au modèle F-Pack Matrix"""
        column_widths = {
            'FPack_Number': 15,
            'Robot_Location_Code': 20,
            'Projet': 25,
            'Client': 20,
            'Sous_Projet': 25,
            'Contractor': 15,
            'Required_Delivery_Time': 18,
            'Delivery_Site': 20,
            'Tracking': 12,
            'FPack_Template': 25,
            'FPack_Abbreviation': 15,
            'Produits_Seuls': 30,  
            'Equipements_Seuls': 30 
        }
        
        for col_idx, column_name in enumerate(df.columns, 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            width = column_widths.get(column_name, 12)
            ws.column_dimensions[column_letter].width = width
        
        ws.row_dimensions[1].height = 30 
        
        ws.freeze_panes = 'A2'
    
    def _auto_adjust_columns(self, ws):
        """Ajuste automatiquement la largeur des colonnes si nécessaire"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)

@router.post("/export/fpack-matrix")
async def export_fpack_matrix(
    request_data: Dict[str, Any], 
    db: Session = Depends(get_db)
):
    """
    Endpoint principal pour l'export de la matrice F-Pack
    
    Request body:
    {
        "fpack_ids": [1, 2, 3, ...] // IDs des SousProjetFpack
    }
    """
    try:
        if "fpack_ids" not in request_data:
            raise HTTPException(
                status_code=400,
                detail="Le champ 'fpack_ids' est requis"
            )
        
        fpack_ids = request_data["fpack_ids"]
        
        if not isinstance(fpack_ids, list) or not fpack_ids:
            raise HTTPException(
                status_code=400,
                detail="'fpack_ids' doit être une liste non vide d'identifiants"
            )
        
        try:
            fpack_ids = [int(fpack_id) for fpack_id in fpack_ids]
        except (ValueError, TypeError):
            raise HTTPException(
                status_code=400,
                detail="Tous les 'fpack_ids' doivent être des entiers valides"
            )
        
        export_service = FPackExportService(db)
        export_data = export_service.get_fpack_data_for_export(fpack_ids)
        
        if not export_data:
            raise HTTPException(
                status_code=404,
                detail="Aucune donnée trouvée pour les instances F-Pack spécifiées"
            )
        
        excel_buffer = export_service.create_excel_export(export_data)
        
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"fpack_matrix_export_{timestamp}.xlsx"
                
        return StreamingResponse(
            BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur interne du serveur: {str(e)}"
        )

@router.get("/export/fpack-matrix/preview/{fpack_id}")
async def preview_fpack_export(fpack_id: int, db: Session = Depends(get_db)):
    """
    Aperçu des données d'une instance F-Pack spécifique pour l'export
    """
    try:
        export_service = FPackExportService(db)
        
        export_data = export_service.get_fpack_data_for_export([fpack_id])
        
        if not export_data:
            raise HTTPException(
                status_code=404,
                detail=f"Instance F-Pack avec l'ID {fpack_id} non trouvée"
            )
        
        return {
            "success": True,
            "fpack_data": export_data[0],
            "total_columns": len(export_data[0]),
            "available_groups": [key for key in export_data[0].keys() 
                               if key not in ['Projet', 'Client', 'Sous_Projet', 'FPack_Number', 
                                            'Robot_Location_Code', 'Contractor', 'Required_Delivery_Time',
                                            'Delivery_Site', 'Tracking', 'FPack_Template', 'FPack_Abbreviation']]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la génération de l'aperçu: {str(e)}"
        )

@router.get("/export/fpack-matrix/stats")
async def get_export_stats(db: Session = Depends(get_db)):
    """
    Statistiques générales sur les données exportables
    """
    try:
        total_fpacks = db.query(models.SousProjetFpack).count()
        
        fpacks_with_selections = db.query(models.SousProjetFpack.id)\
            .join(models.ProjetSelection)\
            .distinct().count()
        
        clients_with_fpacks = db.query(models.Client.id)\
            .join(models.ProjetGlobal)\
            .join(models.SousProjet)\
            .join(models.SousProjetFpack)\
            .distinct().count()
        
        projets_with_fpacks = db.query(models.ProjetGlobal.id)\
            .join(models.SousProjet)\
            .join(models.SousProjetFpack)\
            .distinct().count()
        
        templates_used = db.query(models.FPack.id)\
            .join(models.SousProjetFpack)\
            .distinct().count()
        
        return {
            "success": True,
            "stats": {
                "total_fpack_instances": total_fpacks,
                "fpacks_with_selections": fpacks_with_selections,
                "fpacks_without_selections": total_fpacks - fpacks_with_selections,
                "clients_with_fpacks": clients_with_fpacks,
                "projets_with_fpacks": projets_with_fpacks,
                "templates_used": templates_used,
                "export_readiness": {
                    "ready_for_export": fpacks_with_selections,
                    "percentage": round((fpacks_with_selections / total_fpacks * 100), 2) if total_fpacks > 0 else 0
                }
            }
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la récupération des statistiques: {str(e)}"
        )

@router.post("/export/fpack-matrix/validate")
async def validate_export_request(
    request_data: Dict[str, Any], 
    db: Session = Depends(get_db)
):
    """
    Valide une demande d'export avant exécution
    """
    try:
        
        if "fpack_ids" not in request_data:
            return {
                "valid": False,
                "errors": ["Le champ 'fpack_ids' est requis"]
            }
        
        fpack_ids = request_data["fpack_ids"]
        
        if not isinstance(fpack_ids, list) or not fpack_ids:
            return {
                "valid": False,
                "errors": ["'fpack_ids' doit être une liste non vide"]
            }
        
        try:
            fpack_ids = [int(fpack_id) for fpack_id in fpack_ids]
        except (ValueError, TypeError):
            return {
                "valid": False,
                "errors": ["Tous les 'fpack_ids' doivent être des entiers valides"]
            }
        
        existing_fpacks = db.query(models.SousProjetFpack.id)\
            .filter(models.SousProjetFpack.id.in_(fpack_ids)).all()
        
        existing_ids = {fp.id for fp in existing_fpacks}
        missing_ids = set(fpack_ids) - existing_ids
        
        warnings = []
        if missing_ids:
            warnings.append(f"Instances F-Pack non trouvées: {list(missing_ids)}")
        
        fpacks_with_selections = db.query(models.SousProjetFpack.id)\
            .filter(models.SousProjetFpack.id.in_(list(existing_ids)))\
            .join(models.ProjetSelection)\
            .distinct().all()
        
        fpacks_with_selections_ids = {fp.id for fp in fpacks_with_selections}
        fpacks_without_selections = existing_ids - fpacks_with_selections_ids
        
        if fpacks_without_selections:
            warnings.append(f"Instances F-Pack sans sélections: {list(fpacks_without_selections)}")
        
        return {
            "valid": len(existing_ids) > 0,
            "errors": [],
            "warnings": warnings,
            "summary": {
                "requested_fpacks": len(fpack_ids),
                "existing_fpacks": len(existing_ids),
                "fpacks_with_data": len(fpacks_with_selections_ids),
                "missing_fpacks": len(missing_ids)
            }
        }
        
    except Exception as e:
        return {
            "valid": False,
            "errors": [f"Erreur de validation: {str(e)}"]
        }