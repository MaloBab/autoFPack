from App import models
from App.database import SessionLocal
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
import pandas as pd
import json
import os
from datetime import datetime
from pydantic import BaseModel
from fuzzywuzzy import fuzz
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# Chemin vers le fichier de configuration JSON externe
CONFIG_FILE_PATH = os.path.join(os.path.dirname(__file__), "Column_Config/mapping.json")

# Modèles Pydantic pour la validation des données
class MappingRule(BaseModel):
    target: str
    groupe_nom: Optional[str] = None
    type: Optional[str] = "exact_match"
    search_in: Optional[List[str]] = []
    search_fields: Optional[List[str]] = []

class MappingConfig(BaseModel):
    excel_columns: Dict[str, MappingRule]

class ImportPreviewRequest(BaseModel):
    preview_data: List[Dict[str, Any]]
    mapping_config: MappingConfig
    client_id: int

class ImportExecuteRequest(BaseModel):
    file_data: List[Dict[str, Any]]
    mapping_config: MappingConfig
    client_id: int
    manual_matches: List[Dict[str, Any]]

class ExportRequest(BaseModel):
    project_ids: List[int]
    options: Dict[str, bool]
    format: str = "excel"

class UnmatchedItem(BaseModel):
    id: str
    value: str
    column: str
    suggestions: List[Dict[str, Any]]
    selectedMatch: Optional[Dict[str, Any]] = None

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def load_mapping_config() -> Dict:
    """Charge la configuration de mapping depuis le fichier JSON externe"""
    try:
        if os.path.exists(CONFIG_FILE_PATH):
            with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            # Configuration par défaut basée sur le PDF
            default_config = {
                "excel_columns": {
                    "FPack Number": {
                        "target": "sous_projet_fpack.FPack_number",
                        "type": "direct"
                    },
                    "Plant": {
                        "target": "sous_projet_fpack.plant",
                        "type": "direct"
                    },
                    "Area/Line": {
                        "target": "sous_projet_fpack.area_line",
                        "type": "direct"
                    },
                    "Station/Mode zone": {
                        "target": "sous_projet_fpack.station_mode_zone",
                        "type": "direct"
                    },
                    "Machine code": {
                        "target": "sous_projet_fpack.machine_code",
                        "type": "direct"
                    },
                    "Robot location code": {
                        "target": "sous_projet_fpack.Robot_Location_Code",
                        "type": "direct"
                    },
                    "Area section": {
                        "target": "sous_projet_fpack.area_section",
                        "type": "direct"
                    },
                    "Direct Link": {
                        "target": "sous_projet_fpack.direct_link",
                        "type": "direct"
                    },
                    "Contractor": {
                        "target": "sous_projet_fpack.contractor",
                        "type": "direct"
                    },
                    "Required Delivery time": {
                        "target": "sous_projet_fpack.required_delivery_time",
                        "type": "direct"
                    },
                    "Delivery site": {
                        "target": "sous_projet_fpack.delivery_site",
                        "type": "direct"
                    },
                    "Tracking": {
                        "target": "sous_projet_fpack.tracking",
                        "type": "direct"
                    },
                    "F-pack Type": {
                        "target": "sous_projet_fpack.fpack_type",
                        "type": "direct"
                    },
                    "Mechanical Unit": {
                        "target": "selection",
                        "groupe_nom": "Mechanical Unit",
                        "search_in": ["robots", "equipements"],
                        "search_fields": ["nom", "model"],
                        "type": "fuzzy_match"
                    },
                    "Robot Controller": {
                        "target": "selection",
                        "groupe_nom": "Robot Controller",
                        "search_in": ["produits", "equipements"],
                        "search_fields": ["nom", "description"],
                        "type": "fuzzy_match"
                    },
                    "Media Panel (G)": {
                        "target": "selection",
                        "groupe_nom": "Media Panel",
                        "search_in": ["produits"],
                        "search_fields": ["nom"],
                        "type": "exact_match"
                    },
                    "Key Equipment (G)": {
                        "target": "selection",
                        "groupe_nom": "Key Equipment",
                        "search_in": ["equipements"],
                        "search_fields": ["nom"],
                        "type": "fuzzy_match"
                    }
                }
            }
            save_mapping_config(default_config)
            return default_config
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur configuration : {str(e)}")

def save_mapping_config(config: Dict):
    """Sauvegarde la configuration de mapping dans le fichier JSON externe"""
    try:
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur sauvegarde : {str(e)}")

@router.post("/import/upload")
async def upload_import_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Étape 1 : Upload et analyse initiale du fichier Excel
    """
    try:
        # Vérifier le type de fichier
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Format de fichier non supporté. Utilisez .xlsx ou .xls"
            )
        
        # Lire le contenu binaire du fichier
        content = await file.read()

        # Lire uniquement l’onglet "F-Pack Matrix"
        try:
            df = pd.read_excel(BytesIO(content), sheet_name="F-Pack Matrix", header=4)
        except ValueError:
            # Pandas lève ValueError si l’onglet n’existe pas
            raise HTTPException(
                status_code=400,
                detail="L'onglet 'F-Pack Matrix' est introuvable dans le fichier Excel"
            )
        
        # Nettoyer les colonnes
        df.columns = df.columns.str.strip()
        
        # Supprimer les lignes complètement vides
        df = df.dropna(how='all')
        
        # Prendre les 10 premières lignes pour l'aperçu
        preview_rows = df.head(10).fillna('').to_dict('records')
        
        # Ajouter un status par défaut
        for row in preview_rows:
            row['_status'] = 'pending'
        
        return {
            "success": True,
            "columns": df.columns.tolist(),
            "preview": preview_rows,
            "total_rows": len(df),
            "message": f"Fichier analysé : {len(df)} lignes, {len(df.columns)} colonnes"
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de l'analyse du fichier : {str(e)}")


@router.post("/import/preview")
async def preview_import(
    request: ImportPreviewRequest,
    db: Session = Depends(get_db)
):
    """
    Étape 2 : Aperçu avec mapping et validation
    """
    try:
        processed_data = []
        unmatched_items = []
        summary = {"nb_projets": 0, "nb_sous_projets": 0, "nb_selections": 0}
        
        # Compter les projets uniques
        fpack_numbers = set()
        
        for row_index, row_data in enumerate(request.preview_data):
            processed_row = row_data.copy()
            status = "success"
            errors = []
            
            # Appliquer le mapping pour chaque colonne
            for column, rule in request.mapping_config.excel_columns.items():
                if column not in row_data or not row_data[column]:
                    continue
                
                value = str(row_data[column]).strip()
                
                if rule.target == "selection" and rule.groupe_nom:
                    # Rechercher des correspondances dans les items
                    matches = await find_matching_items(
                        value, 
                        rule.search_in or ["produits", "equipements"], 
                        rule.search_fields or ["nom"],
                        rule.type or "exact_match",
                        db
                    )
                    
                    if not matches:
                        unmatched_items.append({
                            "id": f"{row_index}_{column}",
                            "value": value,
                            "column": column,
                            "suggestions": await get_suggestions_for_value(value, rule.search_in, db)
                        })
                        status = "warning"
                        errors.append(f"Aucune correspondance trouvée pour '{value}' dans {column}")
                    else:
                        summary["nb_selections"] += 1
                
                elif rule.target.startswith("sous_projet_fpack."):
                    # Validation des champs directs
                    field_name = rule.target.split(".")[1]
                    if field_name == "FPack_number" and value:
                        fpack_numbers.add(value)
                    
                    # Validation basique des champs requis
                    if not value and field_name in ["FPack_number"]:
                        status = "error"
                        errors.append(f"{column} est requis")
            
            processed_row['_status'] = status
            processed_row['_errors'] = errors
            processed_data.append(processed_row)
        
        summary["nb_projets"] = len(fpack_numbers)
        summary["nb_sous_projets"] = len(fpack_numbers)        
        return {
            "success": True,
            "processed_data": processed_data,
            "unmatched_items": unmatched_items,
            "summary": summary
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors du traitement : {str(e)}")

@router.post("/import/execute")
async def execute_import(
    request: ImportExecuteRequest,
    db: Session = Depends(get_db)
):
    """
    Étape 3 : Exécution de l'import avec création en base de données
    """
    try:
        results = {"created_projects": 0, "created_selections": 0, "errors": []}
        
        # Transaction pour rollback en cas d'erreur
        db.begin()
        try:
            # 1. Créer le projet global
            projet_global = create_projet_global(
                nom=f"Import_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                client_id=request.client_id,
                db=db
            )
            results["created_projects"] += 1
            
            # 2. Créer les sous-projets et FPacks
            fpacks_created = {}
            for row_data in request.file_data:
                if row_data.get('_status') == 'error':
                    continue
                
                # Créer sous-projet si pas encore créé
                fpack_number = get_mapped_value(row_data, "FPack_number", request.mapping_config)
                if fpack_number and fpack_number not in fpacks_created:
                    sous_projet = create_sous_projet(
                        id_global=projet_global.id,
                        nom=f"Sous-projet_{fpack_number}",
                        db=db
                    )
                    
                    # Créer l'entrée sous_projet_fpack
                    sous_projet_fpack = create_sous_projet_fpack(
                        sous_projet_id=sous_projet.id,
                        fpack_data=extract_fpack_data(row_data, request.mapping_config),
                        db=db
                    )
                    
                    fpacks_created[fpack_number] = sous_projet_fpack
                
                # 3. Créer les sélections
                if fpack_number in fpacks_created:
                    selections_created = create_selections_from_row(
                        row_data, 
                        request.mapping_config, 
                        fpacks_created[fpack_number].id,
                        request.manual_matches,
                        db
                    )
                    results["created_selections"] += len(selections_created)
            
            db.commit()
            
        except Exception as e:
            db.rollback()
            raise e
        
        return {
            "success": True,
            "results": results,
            "message": f"Import terminé : {results['created_projects']} projets, {results['created_selections']} sélections"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'import : {str(e)}")

@router.get("/import/mapping-config")
async def get_mapping_config():
    """
    Récupérer la configuration de mapping depuis le fichier JSON externe
    """
    try:
        config = load_mapping_config()
        return {
            "success": True,
            "config": config
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur récupération config : {str(e)}")

@router.put("/import/mapping-config")
async def update_mapping_config(config: MappingConfig):
    """
    Mettre à jour la configuration de mapping dans le fichier JSON externe
    """
    try:
        # Validation basique
        for column, rule in config.excel_columns.items():
            if rule.target not in ["ignore", "selection"] and not rule.target.startswith("sous_projet_fpack."):
                raise HTTPException(status_code=400, detail=f"Target invalide pour {column}: {rule.target}")
        
        # Sauvegarder dans le fichier JSON
        save_mapping_config(config.dict())
        
        
        return {
            "success": True,
            "message": "Configuration sauvegardée"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur sauvegarde : {str(e)}")

@router.get("/{project_id}/export")
async def export_single_project(
    project_id: int,
    db: Session = Depends(get_db)
):
    """
    Exporter un projet unique avec la structure exacte du PDF
    """
    try:
        project_data = get_project_export_data([project_id], db)
        excel_file = generate_fpack_matrix_export(project_data)
        
        return FileResponse(
            excel_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f'F-Pack_Matrix_{project_id}_{datetime.now().strftime("%d-%m-%Y")}.xlsx'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export : {str(e)}")

@router.post("/export/batch")
async def export_multiple_projects(
    request: ExportRequest,
    db: Session = Depends(get_db)
):
    """
    Exporter plusieurs projets avec la structure F-Pack Matrix
    """
    try:
        project_data = get_project_export_data(request.project_ids, db)
        
        if request.format == "excel":
            excel_file = generate_fpack_matrix_export(project_data, request.options)
            return FileResponse(
                excel_file,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=f'F-Pack_Matrix_{datetime.now().strftime("%d-%m-%Y")}.xlsx'
            )
        else:
            csv_file = generate_csv_export(project_data, request.options)
            return FileResponse(
                csv_file,
                media_type='text/csv',
                filename=f'export_projets_{datetime.now().strftime("%Y%m%d")}.csv'
            )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur export : {str(e)}")

# Fonctions utilitaires améliorées

def generate_fpack_matrix_export(project_data: List[Dict], options: Dict = None) -> str:
    """
    Génère un fichier Excel avec la structure exacte du F-Pack Matrix (PDF)
    """
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "F-Pack Matrix"
        
        # Styles
        header_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ligne 1: Titre du document
        ws.merge_cells('A1:AZ1')
        ws['A1'] = "All non-Italic header columns must be filled in for quotation, except for Comments and Option columns with earlier dependencies. The rest shall be filled in during project."
        
        # Ligne 2: Information de révision
        ws.merge_cells('A2:AZ2')
        ws['A2'] = f"Revision: 3.8 - Exported on {datetime.now().strftime('%d-%m-%Y')}"
        
        # Lignes 3-4: Vides pour correspondre au PDF
        
        # Ligne 5: Headers (exactement comme dans le PDF)
        headers = [
            "FPack Number", "Plant", "Area/Line", "Station/Mode zone", "Machine code",
            "Robot location code", "Area section", "Direct Link", "Comments", "Order Planning",
            "Contractor", "Required Delivery time", "Delivery site", "Tracking",
            "Comments order", "F-pack Type", "Fpack Type", "Fpack abbreviation",
            "Standard or Extended", "Fpack category", "F-pack Comment", "Track motion(7th-axis)option",
            "Comments", "Other Comments", "Mechanical Unit", "Robot", "RCC Length",
            "Cable Teach Length", "Extension for cable Teach", "Retractable TP Cable",
            "Robot base plate", "Transport kit forklift", "Stacking kit", "Power Regeneration",
            "Smooth Stop", "Key Equipment (G)", "Media Panel (G)", "Robot Controller <-> MP (G)",
            "MP <-> Robot (Air) (G)", "MP <-> Equip (Air) (G)", "MP <-> Equip (Water) (G)",
            "MP Support Kit (G)", "Robot Controller <-> Robot (J1) (G)", "Key Equipment (H)",
            "Handling type (H)", "Robot Controller <-> Robot (H)", "Media Panel Air (H)",
            "Robot Controller <-> MP (H)", "MP Support Kit (H)", "MP <-> Robot (H)"
        ]
        
        # Écrire les headers en ligne 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.border = border
            # Rotation du texte pour certaines colonnes comme dans le PDF
            if col > 25:  # Colonnes des équipements
                cell.alignment = Alignment(text_rotation=90, horizontal='center')
        
        # Données des F-Packs (à partir de la ligne 6)
        row_num = 6
        for project in project_data:
            for fpack in project.get('fpacks', []):
                # Écrire les données de base du F-Pack
                ws.cell(row=row_num, column=1, value=fpack.get('FPack_number', ''))
                ws.cell(row=row_num, column=2, value=fpack.get('plant', ''))
                ws.cell(row=row_num, column=3, value=fpack.get('area_line', ''))
                ws.cell(row=row_num, column=4, value=fpack.get('station_mode_zone', ''))
                ws.cell(row=row_num, column=5, value=fpack.get('machine_code', ''))
                ws.cell(row=row_num, column=6, value=fpack.get('Robot_Location_Code', ''))
                ws.cell(row=row_num, column=11, value=fpack.get('contractor', ''))
                ws.cell(row=row_num, column=12, value=fpack.get('required_delivery_time', ''))
                ws.cell(row=row_num, column=13, value=fpack.get('delivery_site', ''))
                ws.cell(row=row_num, column=14, value=fpack.get('tracking', ''))
                
                # Ajouter les sélections dans les colonnes appropriées
                selections = fpack.get('selections', [])
                for selection in selections:
                    groupe_nom = selection.get('groupe_nom', '')
                    item_nom = selection.get('item_nom', '')
                    
                    # Mapper les groupes aux colonnes selon le PDF
                    if groupe_nom == "Mechanical Unit":
                        ws.cell(row=row_num, column=25, value=item_nom)
                    elif groupe_nom == "Robot Controller":
                        ws.cell(row=row_num, column=26, value=item_nom)
                    elif groupe_nom == "Key Equipment":
                        ws.cell(row=row_num, column=36, value=item_nom)
                    elif groupe_nom == "Media Panel":
                        ws.cell(row=row_num, column=37, value=item_nom)
                
                # Appliquer les bordures à toute la ligne
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).border = border
                
                row_num += 1
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        # Ligne de total en bas
        total_row = row_num + 1
        ws.cell(row=total_row, column=1, value="Total")
        ws.cell(row=total_row, column=2, value=len(project_data))
        
        # Informations en bas (comme dans le PDF)
        info_row = total_row + 2
        ws.merge_cells(f'A{info_row}:D{info_row}')
        ws[f'A{info_row}'] = "Dans sous projet fpack"
        
        ws.merge_cells(f'E{info_row}:H{info_row}')
        ws[f'E{info_row}'] = "F-Pack Matrix"
        
        ws.merge_cells(f'I{info_row}:L{info_row}')
        ws[f'I{info_row}'] = "BILA A/S"
        
        ws.merge_cells(f'M{info_row}:P{info_row}')
        ws[f'M{info_row}'] = "Mechanical Unit Generic Handling"
        
        # Sauvegarder
        wb.save(temp_file.name)
        wb.close()
        
        return temp_file.name
        
    except Exception as e:
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise e

async def find_matching_items(value: str, search_in: List[str], search_fields: List[str], match_type: str, db: Session):
    """Trouve des correspondances pour une valeur donnée avec amélioration de performance"""
    matches = []
    
    try:
        for table_name in search_in:
            # Optimisation : utiliser des requêtes ciblées
            if table_name == "produits":
                query = db.query(models.produits)
                if "nom" in search_fields:
                    if match_type == "exact_match":
                        query = query.filter(models.produits.nom.ilike(f'%{value}%'))
                    items = query.limit(50).all()  # Limiter pour performance
            elif table_name == "equipements":
                query = db.query(models.equipements)
                if "nom" in search_fields:
                    if match_type == "exact_match":
                        query = query.filter(models.equipements.nom.ilike(f'%{value}%'))
                    items = query.limit(50).all()
            elif table_name == "robots":
                query = db.query(models.robots)
                if "nom" in search_fields:
                    if match_type == "exact_match":
                        query = query.filter(models.robots.nom.ilike(f'%{value}%'))
                    items = query.limit(50).all()
            else:
                continue
            
            for item in items:
                for field in search_fields:
                    item_value = getattr(item, field, "")
                    if not item_value:
                        continue
                    
                    if match_type == "exact_match":
                        if value.lower() in item_value.lower() or item_value.lower() in value.lower():
                            matches.append({
                                "id": item.id,
                                "nom": item_value,
                                "type": table_name,
                                "score": 100
                            })
                    elif match_type == "fuzzy_match":
                        ratio = fuzz.ratio(value.lower(), item_value.lower())
                        if ratio >= 70:  # Seuil de correspondance ajusté
                            matches.append({
                                "id": item.id,
                                "nom": item_value,
                                "type": table_name,
                                "score": ratio
                            })
        
        # Trier par score décroissant
        matches.sort(key=lambda x: x.get('score', 0), reverse=True)
        return matches[:10] 
        
    except Exception as e:
        return []

async def get_suggestions_for_value(value: str, search_in: List[str], db: Session):
    """Obtient des suggestions pour une valeur non trouvée avec fuzzy matching"""
    suggestions = []
    
    try:
        for table_name in search_in:
            if table_name == "produits":
                items = db.query(models.produits).limit(20).all()
            elif table_name == "equipements":
                items = db.query(models.equipements).limit(20).all()
            elif table_name == "robots":
                items = db.query(models.robots).limit(20).all()
            else:
                continue
                
            for item in items:
                item_name = getattr(item, 'nom', '')
                if item_name:
                    # Calcul du score de similarité
                    score = fuzz.partial_ratio(value.lower(), item_name.lower())
                    suggestions.append({
                        "id": item.id,
                        "nom": item_name,
                        "type": table_name,
                        "score": score
                    })
        
        # Trier par score et retourner les 10 meilleurs
        suggestions.sort(key=lambda x: x['score'], reverse=True)
        return suggestions[:10]
        
    except Exception as e:
        return []

def get_mapped_value(row_data: Dict, field_name: str, mapping_config: MappingConfig) -> str:
    """Extrait une valeur mappée depuis les données de ligne"""
    for column, rule in mapping_config.excel_columns.items():
        if rule.target == f"sous_projet_fpack.{field_name}":
            return str(row_data.get(column, "")).strip()
    return ""

def extract_fpack_data(row_data: Dict, mapping_config: MappingConfig) -> Dict:
    """Extrait les données FPack depuis une ligne"""
    fpack_data = {}
    
    for column, rule in mapping_config.excel_columns.items():
        if rule.target.startswith("sous_projet_fpack."):
            field_name = rule.target.split(".")[1]
            value = str(row_data.get(column, "")).strip()
            if value:  # Ne pas ajouter les valeurs vides
                fpack_data[field_name] = value
    
    return fpack_data

def create_projet_global(nom: str, client_id: int, db: Session):
    """Crée un projet global"""
    try:
        projet = models.ProjetGlobal(
            projet=nom,
            client=client_id
        )
        db.add(projet)
        db.flush()  
        return projet
    except Exception as e:
        raise e

def create_sous_projet(id_global: int, nom: str, db: Session):
    """Crée un sous-projet"""
    try:
        sous_projet = models.SousProjet(
            id_global=id_global,
            nom=nom
        )
        db.add(sous_projet)
        db.flush()
        return sous_projet
    except Exception as e:
        raise e

def create_sous_projet_fpack(sous_projet_id: int, fpack_data: Dict, db: Session):
    """Crée une entrée sous_projet_fpack avec toutes les données du PDF"""
    try:
        # Récupérer ou créer le FPack si nécessaire
        fpack_id = None
        fpack_number = fpack_data.get('FPack_number')
        
        if fpack_number:
            # Chercher si le FPack existe déjà
            existing_fpack = db.query(models.FPack).filter(
                models.FPack.nom == fpack_number
            ).first()
            
            if not existing_fpack:
                # Créer un nouveau FPack
                new_fpack = models.FPack(
                    nom=fpack_number,
                    client=1,  # À adapter selon votre logique
                    fpack_abbr=fpack_number[:10]  # Abbréviation
                )
                db.add(new_fpack)
                db.flush()
                fpack_id = new_fpack.id
            else:
                fpack_id = existing_fpack.id
        
        # Créer l'entrée sous_projet_fpack avec tous les champs du PDF
        sous_projet_fpack = models.SousProjetFpack(
            sous_projet_id=sous_projet_id,
            fpack_id=fpack_id,
            FPack_number=fpack_data.get('FPack_number', ''),
            Robot_Location_Code=fpack_data.get('Robot_Location_Code', ''),
            contractor=fpack_data.get('contractor', ''),
            required_delivery_time=fpack_data.get('required_delivery_time', ''),
            delivery_site=fpack_data.get('delivery_site', ''),
            tracking=fpack_data.get('tracking', ''),
            # Nouveaux champs basés sur le PDF
            plant=fpack_data.get('plant', ''),
            area_line=fpack_data.get('area_line', ''),
            station_mode_zone=fpack_data.get('station_mode_zone', ''),
            machine_code=fpack_data.get('machine_code', ''),
            area_section=fpack_data.get('area_section', ''),
            direct_link=fpack_data.get('direct_link', ''),
            fpack_type=fpack_data.get('fpack_type', ''),
            fpack_category=fpack_data.get('fpack_category', ''),
            track_motion_option=fpack_data.get('track_motion_option', ''),
            standard_or_extended=fpack_data.get('standard_or_extended', ''),
            comments=fpack_data.get('comments', ''),
            other_comments=fpack_data.get('other_comments', '')
        )
        
        db.add(sous_projet_fpack)
        db.flush()
        return sous_projet_fpack
        
    except Exception as e:
        raise e

def create_selections_from_row(row_data: Dict, mapping_config: MappingConfig, 
                              sous_projet_fpack_id: int, manual_matches: List, db: Session):
    """Crée les sélections depuis une ligne de données"""
    selections_created = []
    
    try:
        # Créer un dictionnaire des correspondances manuelles pour accès rapide
        manual_matches_dict = {}
        for match in manual_matches:
            key = f"{match.get('row_index')}_{match.get('column')}"
            manual_matches_dict[key] = match.get('selectedMatch')
        
        for row_index, row_item in enumerate([row_data]):  # Traiter une seule ligne
            for column, rule in mapping_config.excel_columns.items():
                if rule.target != "selection" or not rule.groupe_nom:
                    continue
                
                if column not in row_data or not row_data[column]:
                    continue
                
                value = str(row_data[column]).strip()
                match_key = f"{row_index}_{column}"
                
                # Vérifier s'il y a une correspondance manuelle
                selected_item = manual_matches_dict.get(match_key)
                
                if not selected_item:
                    # Chercher automatiquement
                    matches = find_matching_items(
                        value,
                        rule.search_in or ["produits", "equipements"],
                        rule.search_fields or ["nom"],
                        rule.type or "exact_match",
                        db
                    )
                    if matches:
                        selected_item = matches[0]  # Prendre le meilleur match
                
                if selected_item:
                    # Trouver ou créer le groupe
                    groupe = db.query(models.Groupes).filter(
                        models.Groupes.nom == rule.groupe_nom
                    ).first()
                    
                    if not groupe:
                        groupe = models.Groupes(nom=rule.groupe_nom)
                        db.add(groupe)
                        db.flush()
                    
                    # Créer la sélection
                    selection = models.ProjetSelection(
                        sous_projet_fpack_id=sous_projet_fpack_id,
                        groupe_id=groupe.id,
                        type_item=selected_item['type'],
                        ref_id=selected_item['id']
                    )
                    
                    db.add(selection)
                    selections_created.append(selection)
        
        db.flush()
        return selections_created
        
    except Exception as e:
        raise e

def get_project_export_data(project_ids: List[int], db: Session) -> List[Dict]:
    """Récupère les données des projets pour export avec structure complète"""
    try:
        projects_data = []
        
        for project_id in project_ids:
            # Récupérer le projet global
            projet = db.query(models.ProjetGlobal).filter(
                models.ProjetGlobal.id == project_id
            ).first()
            
            if not projet:
                continue
            
            # Récupérer tous les sous-projets
            sous_projets = db.query(models.SousProjet).filter(
                models.SousProjet.id_global == project_id
            ).all()
            
            project_info = {
                "id": projet.id,
                "nom": projet.projet,
                "client_nom": getattr(projet.client_rel, 'nom', 'N/A') if hasattr(projet, 'client_rel') else 'N/A',
                "fpacks": []
            }
            
            # Pour chaque sous-projet, récupérer les FPacks
            for sous_projet in sous_projets:
                fpacks = db.query(models.SousProjetFpack).filter(
                    models.SousProjetFpack.sous_projet_id == sous_projet.id
                ).all()
                
                for fpack in fpacks:
                    # Récupérer les sélections pour ce FPack
                    selections = db.query(models.ProjetSelection).filter(
                        models.ProjetSelection.sous_projet_fpack_id == fpack.id
                    ).all()
                    
                    selections_data = []
                    for selection in selections:
                        # Récupérer les détails de l'item sélectionné
                        item_data = get_item_details(selection.type_item, selection.ref_id, db)
                        if item_data:
                            selections_data.append({
                                "groupe_id": selection.groupe_id,
                                "groupe_nom": get_groupe_name(selection.groupe_id, db),
                                "type_item": selection.type_item,
                                "item_id": selection.ref_id,
                                "item_nom": item_data.get('nom', ''),
                                "item_details": item_data
                            })
                    
                    fpack_info = {
                        "id": fpack.id,
                        "FPack_number": fpack.FPack_number or '',
                        "Robot_Location_Code": fpack.Robot_Location_Code or '',
                        "contractor": fpack.contractor or '',
                        "required_delivery_time": fpack.required_delivery_time or '',
                        "delivery_site": fpack.delivery_site or '',
                        "tracking": fpack.tracking or '',
                        # Nouveaux champs du PDF
                        "plant": getattr(fpack, 'plant', '') or '',
                        "area_line": getattr(fpack, 'area_line', '') or '',
                        "station_mode_zone": getattr(fpack, 'station_mode_zone', '') or '',
                        "machine_code": getattr(fpack, 'machine_code', '') or '',
                        "area_section": getattr(fpack, 'area_section', '') or '',
                        "direct_link": getattr(fpack, 'direct_link', '') or '',
                        "fpack_type": getattr(fpack, 'fpack_type', '') or '',
                        "fpack_category": getattr(fpack, 'fpack_category', '') or '',
                        "track_motion_option": getattr(fpack, 'track_motion_option', '') or '',
                        "standard_or_extended": getattr(fpack, 'standard_or_extended', '') or '',
                        "comments": getattr(fpack, 'comments', '') or '',
                        "other_comments": getattr(fpack, 'other_comments', '') or '',
                        "selections": selections_data
                    }
                    
                    project_info["fpacks"].append(fpack_info)
            
            projects_data.append(project_info)
        
        return projects_data
        
    except Exception as e:
        raise e

def get_item_details(type_item: str, ref_id: int, db: Session) -> Dict:
    """Récupère les détails d'un item selon son type"""
    try:
        if type_item == "produits":
            item = db.query(models.produits).filter(models.produits.id == ref_id).first()
        elif type_item == "equipements":
            item = db.query(models.equipements).filter(models.equipements.id == ref_id).first()
        elif type_item == "robots":
            item = db.query(models.robots).filter(models.robots.id == ref_id).first()
        else:
            return {}
        
        if item:
            return {
                "id": item.id,
                "nom": getattr(item, 'nom', ''),
                "description": getattr(item, 'description', ''),
                "model": getattr(item, 'model', ''),
                "type": type_item
            }
        
        return {}
        
    except Exception as e:
        return {}

def get_groupe_name(groupe_id: int, db: Session) -> str:
    """Récupère le nom d'un groupe"""
    try:
        groupe = db.query(models.Groupes).filter(
            models.Groupes.id == groupe_id
        ).first()
        return groupe.nom if groupe else ""
    except Exception as e:
        return ""

def generate_csv_export(project_data: List[Dict], options: Dict = None) -> str:
    """Génère un fichier CSV d'export simple"""
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
    
    try:
        # Aplatir les données pour CSV
        rows = []
        for project in project_data:
            for fpack in project.get('fpacks', []):
                row = {
                    'Project_ID': project['id'],
                    'Project_Name': project['nom'],
                    'Client': project['client_nom'],
                    'FPack_Number': fpack['FPack_number'],
                    'Robot_Location_Code': fpack['Robot_Location_Code'],
                    'Contractor': fpack['contractor'],
                    'Required_Delivery_Time': fpack['required_delivery_time'],
                    'Delivery_Site': fpack['delivery_site'],
                    'Tracking': fpack['tracking']
                }
                
                # Ajouter les sélections sous forme de colonnes
                for selection in fpack.get('selections', []):
                    col_name = f"{selection['groupe_nom']}_Selection"
                    row[col_name] = selection['item_nom']
                
                rows.append(row)
        
        df = pd.DataFrame(rows)
        df.to_csv(temp_file.name, index=False, encoding='utf-8')
        
        return temp_file.name
        
    except Exception as e:
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise e