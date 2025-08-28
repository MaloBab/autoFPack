from fastapi import APIRouter, Depends, File, UploadFile, HTTPException # type: ignore
from sqlalchemy import inspect # type: ignore
from sqlalchemy.orm import Session  # type: ignore
from App.database import SessionLocal
from App import models
from App.export_fpack_to_excel import export_fpack_config, export_all_fpacks
from fastapi.responses import StreamingResponse # type: ignore
from io import BytesIO
import os
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill # type: ignore
from reportlab.lib.pagesizes import A4 # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer # type: ignore
import openpyxl # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def validate_and_parse_excel(file: UploadFile, table_name: str, db, required_fields: list[str] = []):
    content = file.file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    inspector = inspect(db.bind)
    use_sql_server = os.getenv("USE_SQL_SERVER", "false").lower() == "true"

    table_mapping = {
        "clients": "FPM_clients",
        "produits": "FPM_produits",
        "equipements": "FPM_equipements",
        "robots": "FPM_robots",
        "fournisseurs": "FPM_fournisseurs",
        "fpacks": "FPM_fpacks",
        "prix": "FPM_prix",
        "prix_robot": "FPM_prix_robot",
        "projets": "FPM_projets_global",
    }
    
    actual_name = table_mapping.get(table_name, table_name)

    expected_columns = [col["name"].lower() for col in inspector.get_columns(actual_name)]
    if "id" in expected_columns:
        expected_columns.remove("id")

    header = [str(cell.value).lower().strip() for cell in next(ws.iter_rows(max_row=1))]
    
    if "fournisseur" in header:
        header[header.index("fournisseur")] = "fournisseur_id"

    idx_id = header.index("id") if "id" in header else None
    if idx_id is not None:
        header.pop(idx_id)

    for col in header:
        if col not in expected_columns:
            raise HTTPException(status_code=400, detail=f"Colonne inconnue dans Excel : {col}")

    data_rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            break 

        row = list(row)
        if idx_id is not None:
            row.pop(idx_id)
        entry = dict(zip(header, row))

        for field in required_fields:
            if entry.get(field) in [None, ""]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Champ '{field}' vide à la ligne {row_num} dans le fichier Excel"
                )

        data_rows.append(entry)

    return data_rows
        
#EXPORT PRODUITS

@router.get("/produits/export/excel")
def export_produits_excel(db: Session = Depends(get_db)):
    produits = db.query(models.Produit).all()
    fournisseurs = [f.nom for f in db.query(models.Fournisseur).all()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    headers = ["Reference", "Nom", "Description", "Fournisseur", "Type"]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for p in produits:
        fournisseur_nom = p.fournisseur.nom if p.fournisseur else ""
        ws.append([
            p.reference or "",
            p.nom or "",
            p.description or "",
            fournisseur_nom,
            p.type or ""
        ])

    dv = DataValidation(type="list", formula1=f'"{",".join(fournisseurs)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D2:D{ws.max_row}")

    for col_idx, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produits.xlsx"},
    )
    
#IMPORT PRODUITS

@router.post("/produits/import/add")
async def import_produits_add(file: UploadFile = File(...), db: Session = Depends(get_db)):
    
    data_rows = validate_and_parse_excel(file, "produits", db, required_fields=["nom", "reference", "fournisseur_id"])  
    fournisseurs_non_trouves = set()
    for row in data_rows:
        fournisseur_nom = row.get('fournisseur_id')
        fournisseur = db.query(models.Fournisseur).filter(models.Fournisseur.nom == fournisseur_nom).first()
        if not fournisseur:
            fournisseurs_non_trouves.add(fournisseur_nom)
            continue

        produit = models.Produit(
            reference=row.get('reference', ''),
            nom=row.get('nom'),
            description=row.get('description', ''),
            fournisseur_id=fournisseur.id,
            type=row.get('type')
        )        
        db.add(produit)

    if fournisseurs_non_trouves:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Fournisseurs non trouvés : {', '.join(str(f) for f in fournisseurs_non_trouves)}")
    db.commit()

    return {"message": f"{len(data_rows)} produits ajoutés"}

#EXPORT ROBOTS

@router.get("/robots/export/excel")
def export_robots_excel(db: Session = Depends(get_db)):
    robots = db.query(models.Robots).all()
    clients = [c.nom for c in db.query(models.Client).all()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Robots"

    headers = ["Reference", "Nom", "Generation", "Client", "Payload", "Range"]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for r in robots:
        client = db.query(models.Client).filter(models.Client.id == r.client).first()
        client_nom = client.nom if client else ""
        ws.append([
            r.reference or "",
            r.nom or "",
            r.generation or "",
            client_nom,
            r.payload or 0,
            r.range or 0
        ])

    dv = DataValidation(type="list", formula1=f'"{",".join(clients)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D2:D{ws.max_row}")

    for col_idx, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=robots.xlsx"},
    )
    
#IMPORT ROBOTS

@router.post("/robots/import/add")
async def import_robots_add(file: UploadFile = File(...), db: Session = Depends(get_db)):
    
    data_rows = validate_and_parse_excel(file, "robots", db, required_fields=["reference", "nom", "generation", "client", "payload", "range"])    
    clients_non_trouves = set()
    for row in data_rows:
        client_nom = row.get('client')
        client = db.query(models.Client).filter(models.Client.nom == client_nom).first()
        if not client:
            clients_non_trouves.add(client_nom)
            continue

        robot = models.Robots(
            reference = row.get('reference', ''),
            nom=row.get('nom'),
            generation=row.get('generation', ''),
            client=client.id,
            payload=row.get('payload', ''),
            range=row.get('range', '')
        )        
        db.add(robot)

    if clients_non_trouves:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Clients non trouvés : {', '.join(str(f) for f in clients_non_trouves)}")
    db.commit()

    return {"message": f"{len(data_rows)} robots ajoutés"}

#EXPORT FPACK TO EXCEL
@router.post("/export-fpack/{fpack_id}")
def export_fpack(fpack_id: int, db: Session = Depends(get_db)):
    wb = export_fpack_config(fpack_id, db)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        'Content-Disposition': f'attachment; filename="F-Pack-{fpack_id}.xlsx"'
    }

    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.get("/export-fpacks/all")
def export_all_fpacks_route(db: Session = Depends(get_db)):
    wb = export_all_fpacks(db)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename="FPacks-All.xlsx"'
    }

    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)