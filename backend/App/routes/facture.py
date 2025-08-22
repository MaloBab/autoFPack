from fastapi import APIRouter, Depends, HTTPException # type: ignore
from fastapi.responses import StreamingResponse# type: ignore
from sqlalchemy.orm import Session # type: ignore
from typing import Dict, Any, List
import io
from datetime import datetime

# Imports pour PDF
from reportlab.lib import colors # type: ignore
from reportlab.lib.pagesizes import A4 # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer # type: ignore
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle # type: ignore
from reportlab.lib.units import cm # type: ignore
from reportlab.lib.enums import TA_CENTER # type: ignore

# Imports pour Excel
import openpyxl # type: ignore
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

from App.database import SessionLocal
from App import models

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_item_label(item_type: str, ref_id: int, db: Session) -> str:
    """Utilitaire pour récupérer le label d'un item selon son type"""
    if item_type == "produit":
        produit = db.query(models.Produit).filter_by(id=ref_id).first()
        return produit.nom if produit else f"(produit {ref_id})"
    elif item_type == "equipement":
        eq = db.query(models.Equipements).filter_by(id=ref_id).first()
        return eq.nom if eq else f"(équipement {ref_id})"
    elif item_type == "robot":
        robot = db.query(models.Robots).filter_by(id=ref_id).first()
        return f"{robot.nom} ({robot.reference})" if robot else f"(robot {ref_id})"
    return f"{item_type} {ref_id}"

def get_fpack_config(fpack_id: int, db: Session) -> List[Dict]:
    """Récupère la configuration des colonnes d'un FPack"""
    columns = db.query(models.FPackConfigColumn)\
        .filter_by(fpack_id=fpack_id)\
        .order_by(models.FPackConfigColumn.ordre)\
        .all()
    
    result = []
    for col in columns:
        entry = {
            "id": col.id,
            "fpack_id": col.fpack_id,
            "type": col.type,
            "ref_id": col.ref_id,
            "ordre": col.ordre
        }
        
        if col.type == "produit":
            produit = db.query(models.Produit).filter_by(id=col.ref_id).first()
            entry["display_name"] = produit.nom if produit else f"(produit {col.ref_id})"
        
        elif col.type == "equipement":
            eq = db.query(models.Equipements).filter_by(id=col.ref_id).first()
            entry["display_name"] = eq.nom if eq else f"(équipement {col.ref_id})"
        
        elif col.type == "group":
            groupe = db.query(models.Groupes).filter_by(id=col.ref_id).first()
            entry["display_name"] = groupe.nom if groupe else f"(groupe {col.ref_id})"
            
            # Récupérer les items du groupe
            if col.ref_id:
                items = db.query(models.GroupeItem).filter_by(group_id=col.ref_id).all()
                entry["group_items"] = [
                    {
                        "id": item.id,
                        "group_id": item.group_id,
                        "type": item.type,
                        "ref_id": item.ref_id,
                        "label": get_item_label(item.type, item.ref_id, db),
                        "statut": item.statut
                    }
                    for item in items
                ]
            else:
                entry["group_items"] = []
        else:
            entry["display_name"] = f"{col.type} {col.ref_id}"
        
        result.append(entry)
    
    return result

def calculate_equipement_price(equipement_id: int, client_id: int, db: Session) -> Dict[str, float]:
    """Calcule le prix d'un équipement basé sur ses produits composants"""
    # Récupérer les produits de l'équipement
    equipement_produits = db.query(models.Equipement_Produit)\
        .filter_by(equipement_id=equipement_id)\
        .all()
    
    total_prix_produit = 0.0
    total_prix_transport = 0.0
    
    for ep in equipement_produits:
        # Récupérer le prix du produit pour ce client
        prix = db.query(models.Prix).filter(
            models.Prix.produit_id == ep.produit_id,
            models.Prix.client_id == client_id
        ).first()
        
        if prix:
            prix_unitaire = float(prix.prix_produit) if prix.prix_produit else 0.0
            prix_transport = float(prix.prix_transport) if prix.prix_transport else 0.0
            
            # Multiplier par la quantité dans l'équipement
            quantite = ep.quantite if ep.quantite else 1
            total_prix_produit += prix_unitaire * quantite
            total_prix_transport += prix_transport * quantite
    
    return {
        "prix_produit": total_prix_produit,
        "prix_transport": total_prix_transport
    }

def get_sous_projet_fpack_facture(sous_projet_fpack_id: int, db: Session):
    """
    Récupère les données de facture pour un sous_projet_fpack donné
    incluant produits seuls, équipements seuls et sélections dans les groupes
    Les items identiques sont regroupés avec quantité cumulée
    """
    # Récupérer le sous_projet_fpack avec ses relations
    sous_projet_fpack = db.query(models.SousProjetFpack).filter(
        models.SousProjetFpack.id == sous_projet_fpack_id
    ).first()
    
    if not sous_projet_fpack:
        raise HTTPException(status_code=404, detail="Sous-projet FPack non trouvé")
    
    # Récupérer le sous-projet
    sous_projet = db.query(models.SousProjet).filter(
        models.SousProjet.id == sous_projet_fpack.sous_projet_id
    ).first()
    
    # Récupérer le projet global
    projet_global = None
    if sous_projet:
        projet_global = db.query(models.ProjetGlobal).filter(
            models.ProjetGlobal.id == sous_projet.id_global
        ).first()
    
    # Récupérer le client via le projet global
    client = None
    if projet_global:
        client = db.query(models.Client).filter(
            models.Client.id == projet_global.client
        ).first()
    
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # Récupérer le FPack
    fpack = db.query(models.FPack).filter(
        models.FPack.id == sous_projet_fpack.fpack_id
    ).first()
    
    # Récupérer la configuration du FPack
    fpack_config = get_fpack_config(sous_projet_fpack.fpack_id, db)
    
    # Dictionnaire pour regrouper les items identiques
    # Clé: (type_item, ref_id), Valeur: données de l'item
    items_grouped = {}
    
    total_produit = 0
    total_transport = 0
    
    # 1. Traiter les éléments de configuration FPack (produits seuls et équipements seuls)
    for config_item in fpack_config:
        if config_item["type"] == "produit":
            # Produit seul dans la configuration
            produit = db.query(models.Produit).filter(
                models.Produit.id == config_item["ref_id"]
            ).first()
            
            if produit:
                key = ("produit", produit.id)
                
                # Récupérer le prix pour ce client
                prix = db.query(models.Prix).filter(
                    models.Prix.produit_id == produit.id,
                    models.Prix.client_id == client.id
                ).first()
                
                prix_unitaire = float(prix.prix_produit) if prix and prix.prix_produit else 0.0
                prix_transport = float(prix.prix_transport) if prix and prix.prix_transport else 0.0
                commentaire = prix.commentaire if prix else None
                
                if key in items_grouped:
                    # Augmenter la quantité seulement
                    items_grouped[key]["qte"] += 1
                else:
                    # Nouvel item
                    items_grouped[key] = {
                        "type": "produit",
                        "produit_id": produit.id,
                        "nom": produit.nom,
                        "qte": 1,
                        "prix_unitaire": prix_unitaire,
                        "prix_transport": prix_transport,
                        "commentaire": commentaire,
                        "total_ligne": 0.0  # Sera calculé à la fin
                    }
        
        elif config_item["type"] == "equipement":
            # Équipement seul dans la configuration
            equipement = db.query(models.Equipements).filter(
                models.Equipements.id == config_item["ref_id"]
            ).first()
            
            if equipement:
                key = ("equipement", equipement.id)
                
                # Calculer le prix de l'équipement
                prix_calc = calculate_equipement_price(equipement.id, client.id, db)
                
                prix_unitaire = prix_calc["prix_produit"]
                prix_transport = prix_calc["prix_transport"]
                
                if key in items_grouped:
                    # Augmenter la quantité seulement
                    items_grouped[key]["qte"] += 1
                else:
                    # Nouvel item
                    items_grouped[key] = {
                        "type": "equipement",
                        "produit_id": equipement.id,
                        "nom": equipement.nom,
                        "qte": 1,
                        "prix_unitaire": prix_unitaire,
                        "prix_transport": prix_transport,
                        "commentaire": "Somme des prix des produits internes",
                        "total_ligne": 0.0  # Sera calculé à la fin
                    }
    
    # 2. Traiter les sélections du projet (dans les groupes)
    selections = db.query(models.ProjetSelection).filter(
        models.ProjetSelection.sous_projet_fpack_id == sous_projet_fpack_id
    ).all()
    
    for selection in selections:
        if selection.type_item == 'produit':
            # Produit sélectionné dans un groupe
            produit = db.query(models.Produit).filter(
                models.Produit.id == selection.ref_id
            ).first()
            
            if produit:
                key = ("produit", produit.id)
                
                prix = db.query(models.Prix).filter(
                    models.Prix.produit_id == produit.id,
                    models.Prix.client_id == client.id
                ).first()
                
                prix_unitaire = float(prix.prix_produit) if prix and prix.prix_produit else 0.0
                prix_transport = float(prix.prix_transport) if prix and prix.prix_transport else 0.0
                commentaire = prix.commentaire if prix else None
                
                if key in items_grouped:
                    # Augmenter la quantité seulement
                    items_grouped[key]["qte"] += 1
                else:
                    # Nouvel item
                    items_grouped[key] = {
                        "type": "produit",
                        "produit_id": produit.id,
                        "nom": produit.nom,
                        "qte": 1,
                        "prix_unitaire": prix_unitaire,
                        "prix_transport": prix_transport,
                        "commentaire": commentaire,
                        "total_ligne": 0.0  # Sera calculé à la fin
                    }
        
        elif selection.type_item == 'robot':
            # Robot sélectionné dans un groupe
            robot = db.query(models.Robots).filter(
                models.Robots.id == selection.ref_id
            ).first()
            
            if robot:
                key = ("robot", robot.id)
                
                prix_robot = db.query(models.PrixRobot).filter(
                    models.PrixRobot.id == robot.id
                ).first()
                
                prix_unitaire = float(prix_robot.prix_robot) if prix_robot and prix_robot.prix_robot else 0.0
                prix_transport = float(prix_robot.prix_transport) if prix_robot and prix_robot.prix_transport else 0.0
                commentaire = prix_robot.commentaire if prix_robot else None
                
                if key in items_grouped:
                    # Augmenter la quantité seulement
                    items_grouped[key]["qte"] += 1
                else:
                    # Nouvel item
                    items_grouped[key] = {
                        "type": "robot",
                        "produit_id": robot.id,
                        "nom": f"{robot.nom} ({robot.reference})",
                        "qte": 1,
                        "prix_unitaire": prix_unitaire,
                        "prix_transport": prix_transport,
                        "commentaire": commentaire,
                        "total_ligne": 0.0  # Sera calculé à la fin
                    }
        
        elif selection.type_item == 'equipement':
            # Équipement sélectionné dans un groupe
            equipement = db.query(models.Equipements).filter(
                models.Equipements.id == selection.ref_id
            ).first()
            
            if equipement:
                key = ("equipement", equipement.id)
                
                # Calculer le prix de l'équipement
                prix_calc = calculate_equipement_price(equipement.id, client.id, db)
                
                prix_unitaire = prix_calc["prix_produit"]
                prix_transport = prix_calc["prix_transport"]
                
                if key in items_grouped:
                    # Augmenter la quantité seulement
                    items_grouped[key]["qte"] += 1
                else:
                    # Nouvel item
                    items_grouped[key] = {
                        "type": "equipement",
                        "produit_id": equipement.id,
                        "nom": equipement.nom,
                        "qte": 1,
                        "prix_unitaire": prix_unitaire,
                        "prix_transport": prix_transport,
                        "commentaire": "Somme des prix des produits internes",
                        "total_ligne": 0.0  # Sera calculé à la fin
                    }
    
    # CORRECTION PRINCIPALE: Convertir le dictionnaire en liste et calculer les totaux APRÈS regroupement
    lines = []
    nb_produits = 0
    nb_robots = 0
    nb_equipements = 0
    
    for item in items_grouped.values():
        # Recalculer le total_ligne basé sur prix unitaires et quantité finale
        item["total_ligne"] = (item["prix_unitaire"] + item["prix_transport"]) * item["qte"]
        lines.append(item)
        
        # Calculer les totaux globaux
        total_produit += item["prix_unitaire"] * item["qte"]
        total_transport += item["prix_transport"] * item["qte"]
        
        # Compter les types d'items
        if item["type"] == "produit":
            nb_produits += item["qte"]
        elif item["type"] == "robot":
            nb_robots += item["qte"]
        elif item["type"] == "equipement":
            nb_equipements += item["qte"]
    
    return {
        "sous_projet_fpack_id": sous_projet_fpack_id,
        "sous_projet_id": sous_projet_fpack.sous_projet_id,
        "nom_sous_projet": sous_projet.nom if sous_projet else "Sous-projet inconnu",
        "projet_global": {
            "id": projet_global.id if projet_global else None,
            "nom": projet_global.projet if projet_global else "Projet inconnu"
        },
        "client_id": client.id,
        "client_nom": client.nom,
        "fpack": {
            "id": fpack.id if fpack else None,
            "nom": fpack.nom if fpack else "FPack inconnu",
            "abbr": fpack.fpack_abbr if fpack else "",
            "FPack_number": sous_projet_fpack.FPack_number,
            "Robot_Location_Code": sous_projet_fpack.Robot_Location_Code
        },
        "currency": "EUR",
        "lines": lines,
        "totaux": {
            "produit": total_produit,
            "transport": total_transport,
            "global": total_produit + total_transport
        },
        "resume": {
            "nb_lignes": len(lines),
            "nb_produits": nb_produits,
            "nb_robots": nb_robots,
            "nb_equipements": nb_equipements
        }
    }
    
    
def get_sous_projet_facture(sous_projet_id: int, db: Session):
    """
    Récupère les données de facture pour un sous-projet donné
    incluant tous les sous_projet_fpack avec leurs configurations complètes
    Les items identiques sont regroupés avec quantité cumulée
    """
    # Récupérer le sous-projet
    sous_projet = db.query(models.SousProjet).filter(
        models.SousProjet.id == sous_projet_id
    ).first()
    
    if not sous_projet:
        raise HTTPException(status_code=404, detail="Sous-projet non trouvé")
    
    # Récupérer le projet global
    projet_global = db.query(models.ProjetGlobal).filter(
        models.ProjetGlobal.id == sous_projet.id_global
    ).first()
    
    # Récupérer le client
    client = None
    if projet_global:
        client = db.query(models.Client).filter(
            models.Client.id == projet_global.client
        ).first()
    
    if not client:
        raise HTTPException(status_code=404, detail="Client non trouvé")
    
    # Récupérer tous les sous_projet_fpack liés
    sous_projets_fpack = db.query(models.SousProjetFpack).filter(
        models.SousProjetFpack.sous_projet_id == sous_projet_id
    ).all()
    
    # Dictionnaire pour regrouper tous les items de tous les FPacks
    items_grouped = {}
    total_produit = 0
    total_transport = 0
    
    for spf in sous_projets_fpack:
        # Récupérer la configuration du FPack
        fpack_config = get_fpack_config(spf.fpack_id, db)
        
        # 1. Traiter les éléments de configuration FPack
        for config_item in fpack_config:
            if config_item["type"] == "produit":
                produit = db.query(models.Produit).filter(
                    models.Produit.id == config_item["ref_id"]
                ).first()
                
                if produit:
                    key = ("produit", produit.id)
                    
                    prix = db.query(models.Prix).filter(
                        models.Prix.produit_id == produit.id,
                        models.Prix.client_id == client.id
                    ).first()
                    
                    prix_unitaire = float(prix.prix_produit) if prix and prix.prix_produit else 0.0
                    prix_transport = float(prix.prix_transport) if prix and prix.prix_transport else 0.0
                    commentaire = prix.commentaire if prix else None
                    
                    if key in items_grouped:
                        items_grouped[key]["qte"] += 1
                        items_grouped[key]["total_ligne"] = (items_grouped[key]["prix_unitaire"] + 
                                                           items_grouped[key]["prix_transport"]) * items_grouped[key]["qte"]
                    else:
                        items_grouped[key] = {
                            "type": "produit",
                            "produit_id": produit.id,
                            "nom": produit.nom,
                            "qte": 1,
                            "prix_unitaire": prix_unitaire,
                            "prix_transport": prix_transport,
                            "commentaire": commentaire,
                            "total_ligne": prix_unitaire + prix_transport
                        }
            
            elif config_item["type"] == "equipement":
                equipement = db.query(models.Equipements).filter(
                    models.Equipements.id == config_item["ref_id"]
                ).first()
                
                if equipement:
                    key = ("equipement", equipement.id)
                    
                    prix_calc = calculate_equipement_price(equipement.id, client.id, db)
                    
                    prix_unitaire = prix_calc["prix_produit"] 
                    prix_transport = prix_calc["prix_transport"]
                    
                    if key in items_grouped:
                        items_grouped[key]["qte"] += 1
                        items_grouped[key]["total_ligne"] = (items_grouped[key]["prix_unitaire"] + 
                                                           items_grouped[key]["prix_transport"]) * items_grouped[key]["qte"]
                    else:
                        items_grouped[key] = {
                            "type": "equipement",
                            "produit_id": equipement.id,
                            "nom": equipement.nom,
                            "qte": 1,
                            "prix_unitaire": prix_unitaire,
                            "prix_transport": prix_transport,
                            "commentaire": "Somme des prix des produits internes",
                            "total_ligne": prix_unitaire + prix_transport
                        }
        
        # 2. Traiter les sélections du projet pour ce FPack
        selections = db.query(models.ProjetSelection).filter(
            models.ProjetSelection.sous_projet_fpack_id == spf.id
        ).all()
        
        for selection in selections:
            if selection.type_item == 'produit':
                produit = db.query(models.Produit).filter(
                    models.Produit.id == selection.ref_id
                ).first()
                
                if produit:
                    key = ("produit", produit.id)
                    
                    prix = db.query(models.Prix).filter(
                        models.Prix.produit_id == produit.id,
                        models.Prix.client_id == client.id
                    ).first()
                    
                    prix_unitaire = float(prix.prix_produit) if prix and prix.prix_produit else 0.0
                    prix_transport = float(prix.prix_transport) if prix and prix.prix_transport else 0.0
                    commentaire = prix.commentaire if prix else None
                    
                    if key in items_grouped:
                        items_grouped[key]["qte"] += 1
                        items_grouped[key]["total_ligne"] = (items_grouped[key]["prix_unitaire"] + 
                                                           items_grouped[key]["prix_transport"]) * items_grouped[key]["qte"]
                    else:
                        items_grouped[key] = {
                            "type": "produit",
                            "produit_id": produit.id,
                            "nom": produit.nom,
                            "qte": 1,
                            "prix_unitaire": prix_unitaire,
                            "prix_transport": prix_transport,
                            "commentaire": commentaire,
                            "total_ligne": prix_unitaire + prix_transport
                        }
            
            elif selection.type_item == 'robot':
                robot = db.query(models.Robots).filter(
                    models.Robots.id == selection.ref_id
                ).first()
                
                if robot:
                    key = ("robot", robot.id)
                    
                    prix_robot = db.query(models.PrixRobot).filter(
                        models.PrixRobot.id == robot.id
                    ).first()
                    
                    prix_unitaire = float(prix_robot.prix_robot) if prix_robot and prix_robot.prix_robot else 0.0
                    prix_transport = float(prix_robot.prix_transport) if prix_robot and prix_robot.prix_transport else 0.0
                    commentaire = prix_robot.commentaire if prix_robot else None
                    
                    if key in items_grouped:
                        items_grouped[key]["qte"] += 1
                        items_grouped[key]["total_ligne"] = (items_grouped[key]["prix_unitaire"] + 
                                                           items_grouped[key]["prix_transport"]) * items_grouped[key]["qte"]
                    else:
                        items_grouped[key] = {
                            "type": "robot",
                            "produit_id": robot.id,
                            "nom": f"{robot.nom} ({robot.reference})",
                            "qte": 1,
                            "prix_unitaire": prix_unitaire,
                            "prix_transport": prix_transport,
                            "commentaire": commentaire,
                            "total_ligne": prix_unitaire + prix_transport
                        }
            
            elif selection.type_item == 'equipement':
                equipement = db.query(models.Equipements).filter(
                    models.Equipements.id == selection.ref_id
                ).first()
                
                if equipement:
                    key = ("equipement", equipement.id)
                    
                    prix_calc = calculate_equipement_price(equipement.id, client.id, db)
                    
                    prix_unitaire = prix_calc["prix_produit"]
                    prix_transport = prix_calc["prix_transport"]
                    
                    if key in items_grouped:
                        items_grouped[key]["qte"] += 1
                        items_grouped[key]["total_ligne"] = (items_grouped[key]["prix_unitaire"] + 
                                                           items_grouped[key]["prix_transport"]) * items_grouped[key]["qte"]
                    else:
                        items_grouped[key] = {
                            "type": "equipement",
                            "produit_id": equipement.id,
                            "nom": equipement.nom,
                            "qte": 1,
                            "prix_unitaire": prix_unitaire,
                            "prix_transport": prix_transport,
                            "commentaire": "Somme des prix des produits internes",
                            "total_ligne": prix_unitaire + prix_transport
                        }
    
    # Convertir le dictionnaire en liste et calculer les totaux
    lines = []
    nb_produits = 0
    nb_robots = 0
    nb_equipements = 0
    
    for item in items_grouped.values():
        lines.append(item)
        
        # Calculer les totaux globaux
        total_produit += item["prix_unitaire"] * item["qte"]
        total_transport += item["prix_transport"] * item["qte"]
        
        # Compter les types d'items
        if item["type"] == "produit":
            nb_produits += item["qte"]
        elif item["type"] == "robot":
            nb_robots += item["qte"]
        elif item["type"] == "equipement":
            nb_equipements += item["qte"]
    
    return {
        "sous_projet_id": sous_projet_id,
        "nom_sous_projet": sous_projet.nom,
        "projet_global": {
            "id": projet_global.id if projet_global else None,
            "nom": projet_global.projet if projet_global else "Projet inconnu"
        },
        "client_id": client.id,
        "client_nom": client.nom,
        "currency": "EUR",
        "lines": lines,
        "totaux": {
            "produit": total_produit,
            "transport": total_transport,
            "global": total_produit + total_transport
        },
        "resume": {
            "nb_lignes": len(lines),
            "nb_produits": nb_produits,
            "nb_robots": nb_robots,
            "nb_equipements": nb_equipements
        }
    }

def generate_pdf_invoice(facture_data: Dict[str, Any]) -> io.BytesIO:
    """Génère une facture PDF esthétique"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, 
                           topMargin=2*cm, bottomMargin=2*cm)
    max_char = 20
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#2563eb'),
        alignment=TA_CENTER
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#1e3a8a')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    story = []
    
    # Titre principal
    story.append(Paragraph("FACTURE", title_style))
    story.append(Spacer(1, 20))
    
    # Informations générales
    info_data = [
        ["Projet Global:", facture_data["projet_global"]["nom"]],
        ["Sous-projet:", facture_data.get("nom_sous_projet", "N/A")],
        ["Client:", facture_data.get("client_nom", "N/A")]
    ]
    
    # Ajouter les infos FPack si disponibles
    if "fpack" in facture_data:
        info_data.extend([
            ["FPack:", f"{facture_data['fpack']['nom']} ({facture_data['fpack']['abbr']})"],
            ["Numéro FPack:", facture_data["fpack"]["FPack_number"] or "N/A"],
            ["Code Robot:", facture_data["fpack"]["Robot_Location_Code"] or "N/A"]
        ])
    
    info_data.extend([
        ["Date:", datetime.now().strftime("%d/%m/%Y")],
        ["Devise:", facture_data["currency"]]
    ])
    
    info_table = Table(info_data, colWidths=[4*cm, 8*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e3a8a')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 30))
    
    # Titre du tableau des produits
    story.append(Paragraph("Détail des éléments", header_style))
    story.append(Spacer(1, 10))
    
    # Tableau des produits
    table_data = [["Élément", "Qté", "Prix €", "Transport €", "Total €", "Commentaire"]]
    
    for line in facture_data["lines"]:
        # Gestion du commentaire tronqué
        raw_comment = line["commentaire"] if line["commentaire"] else ("aucun prix" if line["prix_unitaire"] == 0 else "-")
        if raw_comment and len(str(raw_comment)) > max_char:
            comment = str(raw_comment)[:max_char] + "..."
        else:
            comment = raw_comment
        table_data.append([
            line["nom"],
            str(line["qte"]),
            f"{line['prix_unitaire']:.2f}",
            f"{line['prix_transport']:.2f}",
            f"{line['total_ligne']:.2f}",
            comment
        ])
    
    # Ligne de total
    table_data.append([
        "TOTAL",
        "",
        f"{facture_data['totaux']['produit']:.2f}",
        f"{facture_data['totaux']['transport']:.2f}",
        f"{facture_data['totaux']['global']:.2f} €",
        ""
    ])
    
    # Création du tableau avec largeurs optimisées
    col_widths = [6*cm, 1.5*cm, 2*cm, 2*cm, 2.5*cm, 4*cm]
    products_table = Table(table_data, colWidths=col_widths)
    
    # Style du tableau
    table_style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Corps du tableau
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ALIGN', (1, 1), (4, -2), 'RIGHT'),  # Colonnes numériques à droite
        ('ALIGN', (0, 1), (0, -2), 'LEFT'),   # Colonne produit à gauche
        ('ALIGN', (5, 1), (5, -2), 'LEFT'),   # Colonne commentaire à gauche
        
        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
        
        # Bordures et alternance
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    
    # Mise en évidence des lignes avec prix zéro
    for i, line in enumerate(facture_data["lines"], 1):
        if line["prix_unitaire"] == 0:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fef2f2')))
            table_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.HexColor('#991b1b')))
    
    products_table.setStyle(TableStyle(table_style))
    story.append(products_table)
    
    # Résumé
    story.append(Spacer(1, 30))
    story.append(Paragraph("Résumé", header_style))
    
    resume_data = [
        ["Nombre de lignes:", str(facture_data["resume"]["nb_lignes"])],
        ["Nombre de produits:", str(facture_data["resume"]["nb_produits"])],
        ["Nombre de robots:", str(facture_data["resume"]["nb_robots"])],
        ["Nombre d'équipements:", str(facture_data["resume"]["nb_equipements"])]
    ]
    
    resume_table = Table(resume_data, colWidths=[5*cm, 2*cm])
    resume_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0'))
    ]))
    
    story.append(resume_table)
    
    # Footer
    story.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER
    )
    story.append(Paragraph(f"Facture générée le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", footer_style))
    
    # Construction du PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_invoice(facture_data: Dict[str, Any]) -> io.BytesIO:
    """Génère une facture Excel esthétique"""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture"
    
    # Styles
    title_font = Font(name='Segoe UI', size=18, bold=True, color='2563EB')
    header_font = Font(name='Segoe UI', size=12, bold=True, color='1E3A8A')
    normal_font = Font(name='Segoe UI', size=10)
    bold_font = Font(name='Segoe UI', size=10, bold=True)
    
    # Couleurs
    blue_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    light_blue_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    green_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    red_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    
    # Bordures
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = 'FACTURE'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Informations générales
    row = 3
    info_data = [
        ("Projet Global:", facture_data["projet_global"]["nom"]),
        ("Sous-projet:", facture_data.get("nom_sous_projet", "N/A")),
        ("Client:", facture_data.get("client_nom", "N/A"))
    ]
    
    # Ajouter les infos FPack si disponibles
    if "fpack" in facture_data:
        info_data.extend([
            ("FPack:", f"{facture_data['fpack']['nom']} ({facture_data['fpack']['abbr']})"),
            ("Numéro FPack:", facture_data["fpack"]["FPack_number"] or "N/A"),
            ("Code Robot:", facture_data["fpack"]["Robot_Location_Code"] or "N/A")
        ])
    
    info_data.extend([
        ("Date:", datetime.now().strftime("%d/%m/%Y")),
        ("Devise:", facture_data["currency"])
    ])
    
    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].fill = light_blue_fill
        ws[f'B{row}'].font = normal_font
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    # Espacement
    row += 2
    
    # Headers du tableau
    headers = ["Élément", "Qté", "Prix unitaire €", "Transport unitaire €", "Total €", "Commentaire"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    row += 1
    
    # Données des éléments
    for line in facture_data["lines"]:
        values = [
            line["nom"],
            line["qte"],
            line["prix_unitaire"],
            line["prix_transport"], 
            line["total_ligne"],
            line["commentaire"] if line["commentaire"] else 
            (" Aucun prix" if line["prix_unitaire"] == 0 else "-")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            
            # Format numérique pour les prix
            if col in [2, 3, 4, 5] and isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            
            # Mise en évidence des lignes sans prix
            if line["prix_unitaire"] == 0:
                cell.fill = red_fill
                cell.font = Font(name='Segoe UI', size=10, color='991B1B')
        
        row += 1
    
    # Ligne de total
    total_values = [
        "TOTAL",
        "",
        facture_data["totaux"]["produit"],
        facture_data["totaux"]["transport"],
        f"{facture_data['totaux']['global']} € TTC",
        ""
    ]
    
    for col, value in enumerate(total_values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Ajustement des largeurs de colonnes
    column_widths = [50, 8, 12, 12, 15, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Résumé
    row += 3
    ws[f'A{row}'] = "RÉSUMÉ"
    ws[f'A{row}'].font = header_font
    row += 1
    
    resume_data = [
        ("Nombre de lignes:", facture_data["resume"]["nb_lignes"]),
        ("Nombre de produits:", facture_data["resume"]["nb_produits"]),
        ("Nombre de robots:", facture_data["resume"]["nb_robots"]),
        ("Nombre d'équipements:", facture_data["resume"]["nb_equipements"])
    ]
    
    for label, value in resume_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = normal_font
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    # Footer
    row += 2
    ws[f'A{row}'] = f"Facture générée le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws[f'A{row}'].font = Font(name='Segoe UI', size=8, color='64748B')
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@router.get("/sous_projet_fpack/{sous_projet_fpack_id}/facture-pdf")
async def export_facture_pdf(sous_projet_fpack_id: int, db: Session = Depends(get_db)):
    """Exporte la facture d'un sous-projet FPack en PDF"""
    try:
        facture_data = get_sous_projet_fpack_facture(sous_projet_fpack_id, db)
        pdf_buffer = generate_pdf_invoice(facture_data)
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=facture-projet-{sous_projet_fpack_id}.pdf"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du PDF: {str(e)}")

@router.get("/sous_projet_fpack/{sous_projet_fpack_id}/facture-excel")
async def export_facture_excel(sous_projet_fpack_id: int, db: Session = Depends(get_db)):
    """Exporte la facture d'un sous-projet FPack en Excel"""
    try:
        facture_data = get_sous_projet_fpack_facture(sous_projet_fpack_id, db)
        excel_buffer = generate_excel_invoice(facture_data)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=facture-projet-{sous_projet_fpack_id}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération d'Excel: {str(e)}")

@router.get("/sous_projets/{sous_projet_id}/facture-pdf")
async def export_sous_projet_facture_pdf(sous_projet_id: int, db: Session = Depends(get_db)):
    """Exporte la facture d'un sous-projet en PDF"""
    try:
        facture_data = get_sous_projet_facture(sous_projet_id, db)
        pdf_buffer = generate_pdf_invoice(facture_data)
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=facture-sous-projet-{sous_projet_id}.pdf"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du PDF: {str(e)}")

@router.get("/sous_projets/{sous_projet_id}/facture-excel")
async def export_sous_projet_facture_excel(sous_projet_id: int, db: Session = Depends(get_db)):
    """Exporte la facture d'un sous-projet en Excel"""
    try:
        facture_data = get_sous_projet_facture(sous_projet_id, db)
        excel_buffer = generate_excel_invoice(facture_data)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=facture-sous-projet-{sous_projet_id}.xlsx"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération d'Excel: {str(e)}")

# Endpoint pour obtenir les données de facture en JSON (utilisé par le frontend)
@router.get("/sous_projet_fpack/{sous_projet_fpack_id}/facture")
async def get_facture_sous_projet_fpack(sous_projet_fpack_id: int, db: Session = Depends(get_db)):
    """Récupère les données de facture d'un sous-projet FPack"""
    return get_sous_projet_fpack_facture(sous_projet_fpack_id, db)

@router.get("/sous_projets/{sous_projet_id}/facture")
async def get_facture_sous_projet(sous_projet_id: int, db: Session = Depends(get_db)):
    """Récupère les données de facture d'un sous-projet"""
    return get_sous_projet_facture(sous_projet_id, db)